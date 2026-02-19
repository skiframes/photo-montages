#!/usr/bin/env python3
"""
Edge device Flask application for ski run detection and calibration.
Serves mobile-first calibration UI for coaches to configure trigger zones.
"""

import os
import json
import uuid
import signal
import socket
import tempfile
import threading
from datetime import datetime, timedelta
from pathlib import Path
from subprocess import Popen, PIPE
from typing import Optional

import cv2
import numpy as np
import re
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory

from calibration import (
    detect_gates, compute_calibration, draw_verification_grid,
    draw_gates_on_frame, GATE_SPECS,
)
from stream_manager import StreamManager
from runner import SkiFramesRunner

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

app = Flask(__name__, static_folder='static', template_folder='static')

# Track active processing jobs (subprocess-based: video files)
# Key: job_id, Value: dict with 'process', 'status', 'output', 'config_path', 'video_path'
active_jobs = {}
jobs_lock = threading.Lock()

# Shared RTSP stream manager: one connection per camera, multiple sessions
stream_manager = StreamManager()

# Track in-process RTSP sessions managed by stream_manager
# Key: job_id, Value: dict with 'runner', 'session_id', 'camera_id', 'status', etc.
rtsp_sessions = {}
rtsp_sessions_lock = threading.Lock()

# Configuration
# Use venv python if available (for J40 deployment where packages are in ~/venv/)
_venv_python = Path.home() / 'venv' / 'bin' / 'python3'
PYTHON = str(_venv_python) if _venv_python.exists() else 'python3'

CONFIG_DIR = Path(__file__).parent / 'config' / 'zones'
STITCH_CONFIG_DIR = Path(__file__).parent / 'config' / 'stitch'
CALIBRATION_FRAMES_DIR = Path(tempfile.gettempdir()) / 'skiframes_calibration'
VIDEO_DIR = Path(__file__).parent.parent  # Parent of edge/ for finding test videos
LOGOS_DIR = Path(__file__).parent.parent / 'logos'  # Logo images for overlays
CALIBRATION_CONFIG_DIR = Path(__file__).parent / 'config' / 'calibrations'
UPLOAD_DIR = Path(tempfile.gettempdir()) / 'skiframes_uploads'
CONFIG_DIR.mkdir(parents=True, exist_ok=True)
STITCH_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
CALIBRATION_FRAMES_DIR.mkdir(parents=True, exist_ok=True)
CALIBRATION_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# Track pending gate calibrations (not yet saved to disk)
pending_calibrations = {}
pending_calibrations_lock = threading.Lock()

# Camera definitions with time offset percentages for race timing
# offset_pct: what percentage of run time has elapsed when skier enters this camera's view
CAMERAS = {
    'R1': {
        'name': 'R1',
        'rtsp_url': os.environ.get('R1_RTSP_URL', 'rtsp://j40:J40j40j40@192.168.0.101/h264Preview_01_main'),
        'offset_pct': 0,  # 0% - covers start gate
    },
    'R2': {
        'name': 'R2',
        'rtsp_url': os.environ.get('R2_RTSP_URL', 'rtsp://j40:J40j40j40@192.168.0.102/h264Preview_01_main'),
        'offset_pct': 10,  # 10% into the run
    },
    'Axis': {
        'name': 'Axis',
        'rtsp_url': os.environ.get('AXIS_RTSP_URL', 'rtsp://j40:j40@192.168.1.100/axis-media/media.amp'),
        'offset_pct': 20,  # 20% into the run
    },
    'R3': {
        'name': 'R3',
        'rtsp_url': os.environ.get('R3_RTSP_URL', 'rtsp://j40:J40j40j40@192.168.0.103/h265Preview_01_main'),
        'offset_pct': 50,  # 50% - covers last half of run
    },
}

# Data directory - configurable via environment variable
# Mac: /Volumes/OWC_48/data, Server: /home/pa91/data
DATA_BASE_DIR = Path(os.environ.get('SKIFRAMES_DATA_DIR', '/Volumes/OWC_48/data'))

# Device identity - used in session IDs and for coach page coordination
DEVICE_ID = os.environ.get('SKIFRAMES_DEVICE_ID', socket.gethostname().split('.')[0].lower())

# Admin API URL for device heartbeat
ADMIN_API_URL = os.environ.get('SKIFRAMES_ADMIN_API', 'https://skiframes-admin-api.avillach.workers.dev')

# Vola Excel file location
VOLA_DIR = DATA_BASE_DIR / 'vola'

# Session types and groups
SESSION_TYPES = ['test', 'race', 'gate_training', 'free_skiing']
GROUPS = ['U10', 'U12', 'U14', 'Scored', 'Masters', 'Free Ski']


@app.route('/')
def index():
    """Serve the calibration UI."""
    return send_from_directory('static', 'calibration.html')


@app.route('/api/cameras')
def list_cameras():
    """List available cameras with their time offset percentages."""
    cameras = [
        {
            'id': cid,
            'name': cam['name'],
            'offset_pct': cam.get('offset_pct', 0)
        }
        for cid, cam in CAMERAS.items()
    ]
    return jsonify(cameras)


@app.route('/api/sessions')
def list_sessions():
    """List session types and groups."""
    return jsonify({
        'session_types': SESSION_TYPES,
        'groups': GROUPS
    })


@app.route('/api/vola/files')
def list_vola_files():
    """List available Vola Excel files (searches subdirectories)."""
    if not HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl not installed'}), 500

    files = []
    if VOLA_DIR.exists():
        # Search recursively for xlsx files
        for f in VOLA_DIR.glob('**/*.xlsx'):
            # Show relative path from VOLA_DIR for clarity
            rel_path = f.relative_to(VOLA_DIR)
            files.append({
                'name': str(rel_path),
                'path': str(f),
            })
    return jsonify(sorted(files, key=lambda x: x['name']))


@app.route('/api/vola/startlist-files')
def list_startlist_files():
    """List available start list PDF files (for racer names and teams)."""
    files = []
    if VOLA_DIR.exists():
        # Search recursively for PDF files containing 'start-list'
        for f in VOLA_DIR.glob('**/*start-list*.pdf'):
            rel_path = f.relative_to(VOLA_DIR)
            files.append({
                'name': str(rel_path),
                'path': str(f),
            })
    return jsonify(sorted(files, key=lambda x: x['name']))


def get_gender_from_bib(bib: int) -> str:
    """Determine gender from bib number based on standard ranges."""
    # U12: Women 1-60, Men 61-99
    # U14: Women 111-170, Men 171-220
    if 1 <= bib <= 60 or 111 <= bib <= 170:
        return 'Women'
    elif 61 <= bib <= 110 or 171 <= bib <= 220:
        return 'Men'
    return ''


def parse_startlist_pdf(pdf_path: str) -> dict:
    """
    Parse a start list PDF to extract bib-to-racer mapping.

    Start list format: "111 Stellato Brigid 2013 SUN"
    (bib lastname firstname year team)

    PDF has two columns, so a line might contain two racers.
    Gender is determined by bib number ranges:
    - U12 Women: 1-60, U12 Men: 61-99
    - U14 Women: 111-170, U14 Men: 171-220

    Returns dict mapping bib number to {'name': 'Firstname Lastname', 'team': 'SUN', 'gender': 'Women'}
    """
    if not HAS_PDFPLUMBER:
        return {}

    bib_to_racer = {}

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                for line in text.split('\n'):
                    # Find all racers in the line (may have 2 due to two-column layout)
                    # Pattern: bib (name words) year team
                    # Name can be 2-4 words (e.g., "Van Der Berg John", "Doe Jane", "O'Brien Mary")
                    # Match: bib, then name words, then 4-digit year, then 2-4 letter team code
                    matches = re.findall(r'(\d+)\s+((?:[A-Za-z\'\-]+\s+){1,3}[A-Za-z\'\-]+)\s+(\d{4})\s+([A-Z]{2,4})', line)
                    for match in matches:
                        bib = int(match[0])
                        name_parts = match[1].strip().split()
                        year = match[2]
                        team = match[3]

                        # Name format: last word is firstname, rest is lastname
                        # "Stellato Brigid" -> firstname=Brigid, lastname=Stellato
                        # "Van Der Berg John" -> firstname=John, lastname=Van Der Berg
                        if len(name_parts) >= 2:
                            firstname = name_parts[-1]
                            lastname = ' '.join(name_parts[:-1])  # Keep spaces in multi-word last names
                        else:
                            firstname = name_parts[0] if name_parts else ''
                            lastname = ''

                        bib_to_racer[bib] = {
                            'name': f"{firstname} {lastname}".strip(),
                            'team': team,
                            'gender': get_gender_from_bib(bib)
                        }
    except Exception as e:
        print(f"Error parsing start list PDF {pdf_path}: {e}")

    return bib_to_racer


def parse_results_pdf(pdf_path: str) -> dict:
    """
    Parse a results PDF to extract USSA IDs, rankings, and status (DSQ/DNF/DNS).

    Results format: "1 132 E7031024 Nelson Coulee 2012 PROC 26.89 31.13 58.02"
    (rank bib ussa_id lastname firstname year club run1 run2 total)

    Also parses DSQ/DNF/DNS sections.

    Returns dict mapping bib number to:
    {
        'ussa_id': 'E7031024',
        'name': 'Coulee Nelson',
        'team': 'PROC',
        'gender': 'Women',
        'rank': 1,  # None for DSQ/DNF/DNS
        'status': 'finished',  # or 'DSQ', 'DNF', 'DNS'
        'run1_time': 26.89,  # None for DNS
        'run2_time': 31.13,  # None for DNF in run2
        'total_time': 58.02,  # None for DSQ/DNF/DNS
    }
    """
    if not HAS_PDFPLUMBER:
        return {}

    bib_to_racer = {}
    current_gender = None
    current_status = 'finished'

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                for line in text.split('\n'):
                    # Detect gender sections
                    if 'Women' in line and len(line.strip()) < 20:
                        current_gender = 'Women'
                        continue
                    if 'Men' in line and len(line.strip()) < 20:
                        current_gender = 'Men'
                        continue

                    # Detect status sections
                    if 'Did Not Start' in line:
                        current_status = 'DNS'
                        continue
                    if 'Did Not Finish' in line:
                        current_status = 'DNF'
                        continue
                    if 'Disqualified' in line:
                        current_status = 'DSQ'
                        continue

                    # Parse ranked finishers
                    # Pattern: rank bib [ussa_id] lastname firstname year club run1 run2 total [gap]
                    # Example: "1 132 E7031024 Nelson Coulee 2012 PROC 26.89 31.13 58.02"
                    ranked_match = re.match(
                        r'(\d+)\s+(\d+)\s+(?:(E\d{7})\s+)?((?:[A-Za-z\'\-]+\s+)+)(\d{4})\s+([A-Z]{2,4})\s+([\d:.]+)\s+([\d:.]+)\s+([\d:.]+)',
                        line
                    )
                    if ranked_match:
                        rank = int(ranked_match.group(1))
                        bib = int(ranked_match.group(2))
                        ussa_id = ranked_match.group(3) or ''
                        name_parts = ranked_match.group(4).strip().split()
                        year = ranked_match.group(5)
                        team = ranked_match.group(6)
                        run1 = ranked_match.group(7)
                        run2 = ranked_match.group(8)
                        total = ranked_match.group(9)

                        # Name: last word is firstname
                        if len(name_parts) >= 2:
                            firstname = name_parts[-1]
                            lastname = ' '.join(name_parts[:-1])
                        else:
                            firstname = name_parts[0] if name_parts else ''
                            lastname = ''

                        bib_to_racer[bib] = {
                            'ussa_id': ussa_id,
                            'name': f"{firstname} {lastname}".strip(),
                            'team': team,
                            'gender': current_gender or get_gender_from_bib(bib),
                            'rank': rank,
                            'status': 'finished',
                            'run1_time': parse_race_time(run1),
                            'run2_time': parse_race_time(run2),
                            'total_time': parse_race_time(total),
                        }
                        continue

                    # Parse DSQ/DNF/DNS entries (no rank, may have partial times)
                    # Pattern: bib [ussa_id] lastname firstname year club [times...]
                    # Example: "133 E7090134 Morton Julia 2012 SUN"
                    # Example: "129 E6997139 O'Brien Madison 2012 PROC 32.18"
                    unranked_match = re.match(
                        r'(\d+)\s+(?:(E\d{7})\s+)?((?:[A-Za-z\'\-]+\s+)+)(\d{4})\s+([A-Z]{2,4})(?:\s+([\d:.]+))?',
                        line
                    )
                    if unranked_match and current_status != 'finished':
                        bib = int(unranked_match.group(1))
                        ussa_id = unranked_match.group(2) or ''
                        name_parts = unranked_match.group(3).strip().split()
                        year = unranked_match.group(4)
                        team = unranked_match.group(5)
                        partial_time = unranked_match.group(6)

                        if len(name_parts) >= 2:
                            firstname = name_parts[-1]
                            lastname = ' '.join(name_parts[:-1])
                        else:
                            firstname = name_parts[0] if name_parts else ''
                            lastname = ''

                        bib_to_racer[bib] = {
                            'ussa_id': ussa_id,
                            'name': f"{firstname} {lastname}".strip(),
                            'team': team,
                            'gender': current_gender or get_gender_from_bib(bib),
                            'rank': None,
                            'status': current_status,
                            'run1_time': parse_race_time(partial_time) if partial_time else None,
                            'run2_time': None,
                            'total_time': None,
                        }

        # Fallback: try table extraction for bibs missed by text parsing
        # Some PDF rows overlap with page headers, garbling text extraction
        # but table extraction can still find the cell contents
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row:
                            continue
                        bib = None
                        ussa_id = ''
                        name_cell = ''
                        team = ''
                        for cell in row:
                            if cell is None:
                                continue
                            cell = str(cell).strip()
                            if not bib and re.match(r'^\d{3}$', cell):
                                candidate = int(cell)
                                if 100 <= candidate <= 220 and candidate not in bib_to_racer:
                                    bib = candidate
                            if not bib:
                                bib_in_cell = re.search(r'\b(1[1-9]\d|20\d|21\d|220)\b', cell)
                                if bib_in_cell:
                                    candidate = int(bib_in_cell.group(1))
                                    if candidate not in bib_to_racer:
                                        bib = candidate
                            ussa_match = re.search(r'(E\d{7})', cell)
                            if ussa_match:
                                ussa_id = ussa_match.group(1)
                            name_match = re.search(r'([A-Z][a-z]+(?:[\'\\-][A-Z][a-z]+)?\s+[A-Z][a-z]+)', cell)
                            if name_match and not name_cell:
                                name_cell = name_match.group(1)
                            team_match = re.search(r'\b([A-Z]{2,4})\b', cell)
                            if team_match and team_match.group(1) not in ('None', 'THE', 'AND', 'FOR'):
                                candidate_team = team_match.group(1)
                                if candidate_team in ('SUN', 'PROC', 'RMS', 'FS', 'CMS'):
                                    team = candidate_team

                        if bib and bib not in bib_to_racer and (ussa_id or name_cell):
                            name = ''
                            if name_cell:
                                parts = name_cell.split()
                                if len(parts) >= 2:
                                    name = f"{parts[-1]} {' '.join(parts[:-1])}"
                                else:
                                    name = name_cell
                            bib_to_racer[bib] = {
                                'ussa_id': ussa_id,
                                'name': name,  # Leave empty if not found; startlist can fill it
                                'team': team,
                                'gender': get_gender_from_bib(bib),
                                'rank': None,
                                'status': 'finished',
                                'run1_time': None,
                                'run2_time': None,
                                'total_time': None,
                            }
                            print(f"  Table fallback: Bib {bib} -> {name or '(no name)'} ({team}) ussa={ussa_id}")

    except Exception as e:
        print(f"Error parsing results PDF {pdf_path}: {e}")

    return bib_to_racer


def parse_race_time(time_str: str) -> Optional[float]:
    """Parse race time string to seconds (e.g., '26.89' -> 26.89, '1:03.02' -> 63.02)."""
    if not time_str:
        return None
    try:
        if ':' in time_str:
            parts = time_str.split(':')
            minutes = int(parts[0])
            seconds = float(parts[1])
            return minutes * 60 + seconds
        return float(time_str)
    except (ValueError, IndexError):
        return None


@app.route('/api/vola/results-files')
def list_results_files():
    """List available results PDF files (for racer names) - DEPRECATED, use startlist-files."""
    files = []
    if VOLA_DIR.exists():
        # Search recursively for PDF files containing 'results'
        for f in VOLA_DIR.glob('**/*results*.pdf'):
            rel_path = f.relative_to(VOLA_DIR)
            files.append({
                'name': str(rel_path),
                'path': str(f),
            })
    return jsonify(sorted(files, key=lambda x: x['name']))


@app.route('/api/vola/parse-results', methods=['POST'])
def parse_results():
    """
    Parse a results PDF and return bib-to-name mapping.

    Request body:
    {
        "file_path": "/path/to/results.pdf"
    }
    """
    if not HAS_PDFPLUMBER:
        return jsonify({'error': 'pdfplumber not installed'}), 500

    data = request.get_json() or {}
    file_path = data.get('file_path')

    if not file_path or not Path(file_path).exists():
        return jsonify({'error': 'Invalid file_path'}), 400

    bib_to_name = parse_results_pdf(file_path)

    return jsonify({
        'file': file_path,
        'racers': len(bib_to_name),
        'bib_to_name': bib_to_name,
    })


def parse_vola_time(time_str):
    """
    Parse Vola time string like '10h41:23.9049' or ' 9h52:28.4396' to seconds since midnight.
    Returns None for 'Did Not Start', 'Did Not Finish', 'Disqualified', etc.
    """
    if not time_str or not isinstance(time_str, str):
        return None

    # Strip leading/trailing whitespace
    time_str = time_str.strip()

    # Check for DNF/DNS/DQ
    lower = time_str.lower()
    if 'did not' in lower or 'disqualified' in lower or 'dnf' in lower or 'dns' in lower or 'dq' in lower:
        return None

    # Parse format: 10h41:23.9049 or 9h52:28.4396 (with optional leading space already stripped)
    match = re.match(r'(\d+)h(\d+):(\d+(?:\.\d+)?)', time_str)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        seconds = float(match.group(3))
        return hours * 3600 + minutes * 60 + seconds

    return None


def seconds_to_time_str(seconds):
    """Convert seconds since midnight to HH:MM:SS.mmm format."""
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = seconds % 60
    return f"{hours:02d}:{minutes:02d}:{secs:06.3f}"


@app.route('/api/vola/parse', methods=['POST'])
def parse_vola_file():
    """
    Parse a Vola Excel file and return timing data for a specific race.

    Request body:
    {
        "file_path": "/path/to/vola.xlsx",
        "race": "U12 run 1",  // One of: "U12 run 1", "U12 run 2", "U14 run 1", "U14 run 2"
        "camera_id": "R1",  // Camera to use for offset calculation
    }

    Returns list of racers with their timing windows for the specified camera.
    """
    if not HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl not installed'}), 500

    data = request.get_json() or {}
    file_path = data.get('file_path')
    race = data.get('race', '').lower()
    camera_id = data.get('camera_id', 'R1')

    if not file_path or not Path(file_path).exists():
        return jsonify({'error': 'Invalid file_path'}), 400

    # Map race selection to sheet names
    race_map = {
        'u12 run 1': ('u12 start run 1', 'u12 end run 1'),
        'u12 run 2': ('u12 start run 2', 'u12 end run 2'),
        'u14 run 1': ('u14 start run 1', 'u14 end run 1'),
        'u14 run 2': ('u14 start run 2', 'u14 end run 2'),
    }

    if race not in race_map:
        return jsonify({'error': f'Invalid race. Must be one of: {list(race_map.keys())}'}), 400

    start_sheet, end_sheet = race_map[race]
    camera_offset_pct = CAMERAS.get(camera_id, {}).get('offset_pct', 0) / 100.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Parse start times (Column B = bib, Column C = time)
        start_times = {}
        ws_start = wb[start_sheet]
        for row in ws_start.iter_rows(min_row=2, values_only=True):
            bib_raw = row[1]  # Column B (index 1)
            time_str = row[2]  # Column C (index 2)
            if bib_raw and time_str:
                try:
                    bib = int(bib_raw)  # Convert to int for consistent sorting
                except (ValueError, TypeError):
                    continue
                start_seconds = parse_vola_time(str(time_str))
                if start_seconds is not None:
                    start_times[bib] = start_seconds

        # Parse end times and run durations (Column B = bib, Column C = time, Column D = duration)
        end_times = {}
        run_durations = {}
        ws_end = wb[end_sheet]
        for row in ws_end.iter_rows(min_row=2, values_only=True):
            bib_raw = row[1]  # Column B
            time_str = row[2]  # Column C
            duration = row[3] if len(row) > 3 else None  # Column D
            if bib_raw:
                try:
                    bib = int(bib_raw)  # Convert to int for consistent sorting
                except (ValueError, TypeError):
                    continue
                if time_str:
                    end_seconds = parse_vola_time(str(time_str))
                    if end_seconds is not None:
                        end_times[bib] = end_seconds
                if duration is not None and isinstance(duration, (int, float)):
                    run_durations[bib] = float(duration)

        wb.close()

        # Build racer list with timing windows for the camera
        racers = []
        for bib in sorted(start_times.keys()):
            start_sec = start_times[bib]

            # Calculate run duration from end time or direct duration
            run_duration = None
            if bib in end_times:
                run_duration = end_times[bib] - start_sec
            elif bib in run_durations:
                run_duration = run_durations[bib]

            # Calculate camera-specific timing window
            # Camera offset: when the skier enters this camera's view
            if run_duration and run_duration > 0:
                camera_start_sec = start_sec + (run_duration * camera_offset_pct)
                # Assume camera covers ~20% of the run (configurable)
                camera_duration = run_duration * 0.2
                camera_end_sec = camera_start_sec + camera_duration
            else:
                # No finish time - estimate based on average run duration (~40 seconds)
                estimated_duration = 40.0
                camera_start_sec = start_sec + (estimated_duration * camera_offset_pct)
                camera_duration = estimated_duration * 0.2
                camera_end_sec = camera_start_sec + camera_duration
                run_duration = None  # Mark as estimated

            racer = {
                'bib': bib,
                'start_time_sec': start_sec,
                'start_time_str': seconds_to_time_str(start_sec),
                'run_duration': run_duration,
                'finished': bib in end_times,
                'camera_start_sec': camera_start_sec,
                'camera_start_str': seconds_to_time_str(camera_start_sec),
                'camera_end_sec': camera_end_sec,
                'camera_end_str': seconds_to_time_str(camera_end_sec),
            }
            racers.append(racer)

        # Sort by start time
        racers.sort(key=lambda r: r['start_time_sec'])

        return jsonify({
            'race': race,
            'camera_id': camera_id,
            'camera_offset_pct': camera_offset_pct * 100,
            'racers': racers,
            'total_racers': len(racers),
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/vola/generate-filename', methods=['POST'])
def generate_vola_filename():
    """
    Generate a filename for a photo montage based on Vola data.

    Format: Name_bib#_RunX_CamXX_Ragged_YYYY-Mon-DD.jpg
    Example: JohnDoe_106_R1_CamR1_Ragged_2026-Feb-01.jpg

    Request body:
    {
        "bib": 106,
        "name": "John Doe",  // Optional - will use bib if not provided
        "race": "U12 run 1",
        "camera_id": "R1",
        "date": "2026-02-01"  // Optional - defaults to today
    }
    """
    data = request.get_json() or {}
    bib = data.get('bib')
    name = data.get('name', '').strip()
    race = data.get('race', '').lower()
    camera_id = data.get('camera_id', 'R1')
    date_str = data.get('date')

    if not bib:
        return jsonify({'error': 'bib is required'}), 400

    # Parse race to get run number
    run_num = '1'
    if 'run 2' in race:
        run_num = '2'

    # Format name (remove spaces, use bib if no name)
    if name:
        name_part = name.replace(' ', '')
    else:
        name_part = f"Bib{bib}"

    # Format date
    if date_str:
        try:
            dt = datetime.fromisoformat(date_str)
        except:
            dt = datetime.now()
    else:
        dt = datetime.now()
    date_part = dt.strftime('%Y%b%d')  # e.g., 2026Feb01

    filename = f"{name_part}_{bib}_R{run_num}_Cam{camera_id}_Ragged_{date_part}.jpg"

    return jsonify({
        'filename': filename,
        'bib': bib,
        'name': name_part,
        'run': run_num,
        'camera': camera_id,
        'date': date_part,
    })


@app.route('/api/vola/racers-for-video', methods=['POST'])
def get_racers_for_video():
    """
    Get list of racers whose timing windows fall within a video's timespan.

    This helps identify which racers should be in the montages generated
    from a specific video file.

    Request body:
    {
        "file_path": "/path/to/vola.xlsx",
        "race": "U12 run 1",
        "camera_id": "R1",
        "video_start_time": "10:41:00",  // HH:MM:SS format (Boston time)
        "video_duration_sec": 120  // Video length in seconds
    }

    Returns list of racers with their camera timing windows.
    """
    if not HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl not installed'}), 500

    data = request.get_json() or {}
    file_path = data.get('file_path')
    race = data.get('race', '').lower()
    camera_id = data.get('camera_id', 'R1')
    video_start = data.get('video_start_time', '')
    video_duration = data.get('video_duration_sec', 0)

    if not file_path or not Path(file_path).exists():
        return jsonify({'error': 'Invalid file_path'}), 400

    # Parse video start time
    try:
        parts = video_start.split(':')
        video_start_sec = int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
    except:
        return jsonify({'error': 'Invalid video_start_time format (use HH:MM:SS)'}), 400

    video_end_sec = video_start_sec + video_duration

    # Get all racers for this race
    from flask import url_for
    response = parse_vola_file()
    if response.status_code != 200:
        return response

    all_racers = response.get_json().get('racers', [])

    # Filter to racers whose camera window overlaps with video timespan
    matching_racers = []
    for racer in all_racers:
        cam_start = racer['camera_start_sec']
        cam_end = racer['camera_end_sec']

        # Check if camera window overlaps with video
        if cam_start <= video_end_sec and cam_end >= video_start_sec:
            # Calculate video offset (when racer appears in video)
            video_offset_sec = max(0, cam_start - video_start_sec)
            racer['video_offset_sec'] = video_offset_sec
            racer['video_offset_str'] = f"{int(video_offset_sec // 60)}:{video_offset_sec % 60:05.2f}"
            matching_racers.append(racer)

    return jsonify({
        'race': race,
        'camera_id': camera_id,
        'video_start': video_start,
        'video_duration_sec': video_duration,
        'matching_racers': matching_racers,
        'total_matching': len(matching_racers),
    })


@app.route('/api/frame/grab', methods=['POST'])
def grab_frame():
    """
    Grab a frame from RTSP stream or video file.
    Returns the frame as a JPEG with a unique ID for reference.
    """
    data = request.get_json() or {}
    source_type = data.get('source_type', 'rtsp')  # 'rtsp' or 'video'
    camera_id = data.get('camera_id')
    video_path = data.get('video_path')
    seek_seconds = data.get('seek_seconds', 0)

    if source_type == 'rtsp':
        if not camera_id or camera_id not in CAMERAS:
            return jsonify({'error': 'Invalid camera_id'}), 400
        source = CAMERAS[camera_id]['rtsp_url']
    else:
        if not video_path or not os.path.exists(video_path):
            return jsonify({'error': 'Invalid video_path'}), 400
        source = video_path

    try:
        if source_type == 'rtsp':
            # Use same ffmpeg options as detection engine:
            # - TCP transport for reliable delivery
            # - genpts to fix non-monotonic timestamps (Reolink)
            os.environ["OPENCV_FFMPEG_CAPTURE_OPTIONS"] = "rtsp_transport;tcp|fflags;+genpts|stimeout;5000000"
            cap = cv2.VideoCapture(source, cv2.CAP_FFMPEG)
        else:
            cap = cv2.VideoCapture(source)

        if not cap.isOpened():
            return jsonify({'error': f'Cannot connect to camera stream'}), 500

        if source_type == 'video' and seek_seconds > 0:
            fps = cap.get(cv2.CAP_PROP_FPS)
            if fps > 0:
                cap.set(cv2.CAP_PROP_POS_FRAMES, int(seek_seconds * fps))

        ret, frame = cap.read()
        cap.release()

        if not ret:
            return jsonify({'error': 'Connected but failed to read frame'}), 500

        # Generate unique ID for this frame
        frame_id = str(uuid.uuid4())[:8]
        frame_path = CALIBRATION_FRAMES_DIR / f'{frame_id}.jpg'

        # Save frame
        cv2.imwrite(str(frame_path), frame, [cv2.IMWRITE_JPEG_QUALITY, 90])

        height, width = frame.shape[:2]

        return jsonify({
            'frame_id': frame_id,
            'width': width,
            'height': height,
            'url': f'/api/frame/{frame_id}'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/frame/<frame_id>')
def get_frame(frame_id):
    """Serve a grabbed frame."""
    frame_path = CALIBRATION_FRAMES_DIR / f'{frame_id}.jpg'
    if not frame_path.exists():
        return jsonify({'error': 'Frame not found'}), 404
    return send_file(frame_path, mimetype='image/jpeg')


@app.route('/api/video/info', methods=['POST'])
def video_info():
    """Get video file info for scrubbing."""
    data = request.get_json() or {}
    video_path = data.get('video_path')

    if not video_path or not os.path.exists(video_path):
        return jsonify({'error': 'Invalid video_path'}), 400

    try:
        cap = cv2.VideoCapture(video_path)
        info = {
            'fps': cap.get(cv2.CAP_PROP_FPS),
            'frames': int(cap.get(cv2.CAP_PROP_FRAME_COUNT)),
            'width': int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)),
            'height': int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)),
        }
        info['duration'] = info['frames'] / info['fps'] if info['fps'] > 0 else 0
        cap.release()
        return jsonify(info)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/config/save', methods=['POST'])
def save_config():
    """Save session configuration with trigger zones."""
    data = request.get_json()

    # start_zone is always required
    # end_zone is required unless run_duration_seconds is set (duration mode)
    required = ['camera_id', 'session_type', 'group', 'start_zone']
    for field in required:
        if field not in data:
            return jsonify({'error': f'Missing required field: {field}'}), 400

    # Validate: need either end_zone or run_duration_seconds
    if not data.get('end_zone') and not data.get('run_duration_seconds'):
        return jsonify({'error': 'Either end_zone or run_duration_seconds is required'}), 400

    # Generate session ID: date_time_group_type_camera_device
    # Uses seconds precision to allow multiple simultaneous sessions
    now = datetime.now()
    camera_id = data['camera_id']
    session_id = f"{now.strftime('%Y-%m-%d_%H%M%S')}_{data['group'].lower()}_{data['session_type']}_{camera_id}_{DEVICE_ID}"

    # Default session end time: 90 minutes from now
    session_end = data.get('session_end_time')
    if not session_end:
        session_end = (now + timedelta(minutes=90)).isoformat()

    config = {
        'session_id': session_id,
        'device_id': DEVICE_ID,
        'session_type': data['session_type'],
        'group': data['group'],
        'camera_id': data['camera_id'],
        'camera_url': CAMERAS.get(data['camera_id'], {}).get('rtsp_url', ''),
        'calibration_frame': data.get('calibration_frame_id', ''),
        'start_zone': data['start_zone'],  # {x, y, w, h}
        'end_zone': data.get('end_zone'),  # {x, y, w, h} - optional if using duration mode
        'crop_zone': data.get('crop_zone'),  # {x, y, w, h} - optional, overrides zone-based crop
        'run_duration_seconds': data.get('run_duration_seconds'),  # if set, use duration mode instead of END zone
        'session_start_time': now.isoformat(),
        'session_end_time': session_end,
        'pre_buffer_seconds': data.get('pre_buffer_seconds', 2),
        'post_buffer_seconds': data.get('post_buffer_seconds', 2),
        'detection_threshold': data.get('detection_threshold', 25),
        'min_pixel_change_pct': data.get('min_pixel_change_pct', 5.0),
        'min_brightness': data.get('min_brightness', 100),  # Shadow filter: min brightness of motion pixels
        'discipline': data.get('discipline', 'freeski'),  # freeski, sl_youth, sl_adult, gs_panel, sg_panel
        'montage_fps_list': data.get('montage_fps_list', [4.0]),
        'start_offset_sec': data.get('start_offset_sec', 0),  # Fixed delay in seconds after trigger
        # Vola racer data for naming montages
        'vola_racers': data.get('vola_racers', []),
        'vola_videos': data.get('vola_videos', []),  # Videos to process
        'vola_race': data.get('vola_race', ''),
        'vola_camera': data.get('vola_camera', ''),
        'vola_view': data.get('vola_view', '1'),  # View number 1-5 for filename
        'race_date': data.get('race_date', now.strftime('%Y-%m-%d')),
        'calibration_id': data.get('calibration_id'),  # Links to gate calibration if present
    }

    # Save config
    config_path = CONFIG_DIR / f'{session_id}.json'
    with open(config_path, 'w') as f:
        json.dump(config, f, indent=2)

    return jsonify({
        'success': True,
        'session_id': session_id,
        'config_path': str(config_path)
    })


def parse_iso_datetime(s):
    """Parse ISO datetime string, handling 'Z' suffix for UTC."""
    if s.endswith('Z'):
        s = s[:-1] + '+00:00'
    return datetime.fromisoformat(s)


@app.route('/api/config/active')
def get_active_config():
    """Get the currently active session config (most recent)."""
    configs = sorted(CONFIG_DIR.glob('*.json'), key=lambda p: p.stat().st_mtime, reverse=True)
    if not configs:
        return jsonify({'active': False})

    with open(configs[0]) as f:
        config = json.load(f)

    # Check if session is still active
    try:
        end_time = parse_iso_datetime(config['session_end_time'])
        # Compare as naive datetime (ignore timezone for simplicity)
        if end_time.replace(tzinfo=None) < datetime.now():
            return jsonify({'active': False, 'last_session': config})
    except (ValueError, KeyError):
        return jsonify({'active': False, 'last_session': config})

    return jsonify({'active': True, 'config': config})


@app.route('/api/config/all_active')
def get_all_active_configs():
    """Get ALL sessions that are still active (end_time in future) or have running processes."""
    configs = sorted(CONFIG_DIR.glob('*.json'), key=lambda p: p.stat().st_mtime, reverse=True)
    now = datetime.now()
    sessions = []

    # Collect sessions with running jobs (even if config says expired)
    running_session_ids = {}  # session_id -> job_id mapping

    # Check subprocess jobs
    with jobs_lock:
        for job_id, job in active_jobs.items():
            if job['status'] == 'running':
                # Extract session_id from config_path
                cp = job.get('config_path', '')
                if cp:
                    sid = Path(cp).stem
                    running_session_ids[sid] = job_id

    # Check in-process RTSP sessions (stream manager)
    with rtsp_sessions_lock:
        for job_id, session in rtsp_sessions.items():
            if session['status'] == 'running':
                running_session_ids[session['session_id']] = job_id

    for config_path in configs:
        try:
            with open(config_path) as f:
                config = json.load(f)

            session_id = config_path.stem
            is_running = session_id in running_session_ids

            # Check if session end_time is in the future
            is_active = False
            try:
                end_time = parse_iso_datetime(config['session_end_time'])
                if end_time.replace(tzinfo=None) > now:
                    is_active = True
            except (ValueError, KeyError):
                pass

            if is_active or is_running:
                session_info = {
                    'session_id': session_id,
                    'config': config,
                    'is_running': is_running,
                    'is_active': is_active,
                }
                # Include job_id so frontend can poll metrics/status
                if session_id in running_session_ids:
                    session_info['job_id'] = running_session_ids[session_id]
                sessions.append(session_info)
        except Exception:
            continue

    return jsonify({'sessions': sessions})


@app.route('/api/detection/metrics/<job_id>')
def get_detection_metrics(job_id):
    """Get live detection metrics for a running job (for calibration chart)."""
    with jobs_lock:
        job = active_jobs.get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404

        config_path = job.get('config_path', '')

    if not config_path:
        return jsonify({'error': 'No config path for job'}), 404

    # Derive session output dir from config
    try:
        with open(config_path) as f:
            config = json.load(f)
        session_id = config.get('session_id', Path(config_path).stem)
    except Exception:
        session_id = Path(config_path).stem

    # Look for metrics file in output directory
    output_dir = str(Path(__file__).resolve().parent.parent / 'output')
    metrics_path = os.path.join(output_dir, session_id, 'detection_metrics.json')

    if not os.path.exists(metrics_path):
        return jsonify({'entries': [], 'message': 'No metrics yet'})

    try:
        with open(metrics_path) as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/config/<session_id>')
def get_config(session_id):
    """Get a specific session config."""
    config_path = CONFIG_DIR / f'{session_id}.json'
    if not config_path.exists():
        return jsonify({'error': 'Config not found'}), 404

    with open(config_path) as f:
        config = json.load(f)
    return jsonify(config)


@app.route('/api/config/stop/<session_id>', methods=['POST'])
def stop_session(session_id):
    """Stop a session: set end time to now AND kill any running detection process."""
    config_path = CONFIG_DIR / f'{session_id}.json'
    if not config_path.exists():
        return jsonify({'error': 'Config not found'}), 404

    with open(config_path) as f:
        config = json.load(f)

    # Set end time to now to mark session as ended
    config['session_end_time'] = datetime.now().isoformat()

    with open(config_path, 'w') as f:
        json.dump(config, f, indent=2)

    # Also kill any running process/session for this session
    process_killed = False

    # Check in-process RTSP sessions first
    with rtsp_sessions_lock:
        for job_id, session in rtsp_sessions.items():
            if session['session_id'] == session_id and session['status'] == 'running':
                stream_manager.stop_session(session_id)
                session['status'] = 'stopped'
                process_killed = True
                print(f"[STOP] Stopped in-process RTSP session {session_id} (job {job_id})")
                break

    # Check subprocess jobs
    if not process_killed:
        with jobs_lock:
            for job_id, job in active_jobs.items():
                if job.get('config_path', '').endswith(f'{session_id}.json') and job['status'] == 'running':
                    process = job['process']
                    try:
                        process.terminate()
                        try:
                            process.wait(timeout=3)
                        except:
                            process.kill()
                        job['status'] = 'stopped'
                        process_killed = True
                        print(f"[STOP] Killed process for session {session_id} (job {job_id})")
                    except Exception as e:
                        print(f"[STOP] Failed to kill process for {session_id}: {e}")
                    break

    return jsonify({'success': True, 'session_id': session_id, 'process_killed': process_killed})


@app.route('/api/config/update_live/<session_id>', methods=['POST'])
def update_live_settings(session_id):
    """
    Update detection settings on a running session.
    The detection engine hot-reloads from the config file every ~2 seconds,
    so changes take effect within a few seconds without restarting.
    """
    config_path = CONFIG_DIR / f'{session_id}.json'
    if not config_path.exists():
        return jsonify({'error': 'Config not found'}), 404

    data = request.get_json()

    with open(config_path) as f:
        config = json.load(f)

    # Only allow updating tunable parameters (not zones or session structure)
    updatable = ['detection_threshold', 'min_pixel_change_pct', 'min_brightness',
                 'start_offset_sec', 'session_end_time']
    changed = []
    for key in updatable:
        if key in data:
            old_val = config.get(key)
            config[key] = data[key]
            if old_val != data[key]:
                changed.append(f"{key}: {old_val} -> {data[key]}")

    if changed:
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=2)
        print(f"[LIVE] Updated settings for {session_id}: {', '.join(changed)}")

    return jsonify({
        'success': True,
        'session_id': session_id,
        'changed': changed,
        'detection_threshold': config.get('detection_threshold'),
        'min_pixel_change_pct': config.get('min_pixel_change_pct'),
        'min_brightness': config.get('min_brightness'),
    })


@app.route('/api/videos')
def list_videos():
    """List available video files for processing."""
    videos = []

    # Directories to scan for videos
    video_dirs = [
        VIDEO_DIR,  # Parent of edge/ for finding test videos
        VIDEO_DIR / 'videos-to-process',  # Dedicated folder for videos to process
    ]

    for video_dir in video_dirs:
        if not video_dir.exists():
            continue
        for ext in ['*.mp4', '*.mkv', '*.avi']:
            for video_path in video_dir.glob(ext):
                try:
                    cap = cv2.VideoCapture(str(video_path))
                    fps = cap.get(cv2.CAP_PROP_FPS)
                    if fps <= 0:
                        cap.release()
                        continue
                    duration = cap.get(cv2.CAP_PROP_FRAME_COUNT) / fps
                    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                    cap.release()
                    videos.append({
                        'name': video_path.name,
                        'path': str(video_path),
                        'duration': round(duration, 1),
                        'resolution': f"{width}x{height}",
                        'size_mb': round(video_path.stat().st_size / 1024 / 1024, 1),
                    })
                except:
                    pass

    # Remove duplicates (by path) and sort by name
    seen_paths = set()
    unique_videos = []
    for v in videos:
        if v['path'] not in seen_paths:
            seen_paths.add(v['path'])
            unique_videos.append(v)

    return jsonify(sorted(unique_videos, key=lambda v: v['name']))


# External recordings drive path (uses DATA_BASE_DIR from environment)
RECORDINGS_DIR = DATA_BASE_DIR / 'recordings'

# J40 SD card data directory (flat structure: R1/R1_YYYYMMDD_HHMMSS.mp4)
J40_SD_DATA_DIR = Path(os.environ.get('J40_SD_DATA_DIR', '/Users/paul2/j40/data/sd_data'))


def parse_reolink_video_times(filename: str) -> tuple:
    """
    Parse start and end times from Reolink video filename.
    Format: RecM03_YYYYMMDD_HHMMSS_HHMMSS_XXXXX_XXXXX.mp4

    Returns (date_str, start_time_sec, end_time_sec) or (None, None, None) if parse fails.
    """
    # Example: RecM03_20260201_081538_082039_533C810_BEF5D90.mp4
    match = re.match(r'RecM03_(\d{8})_(\d{6})_(\d{6})_', filename)
    if match:
        date_str = match.group(1)  # YYYYMMDD
        start_str = match.group(2)  # HHMMSS
        end_str = match.group(3)  # HHMMSS

        # Convert to seconds since midnight
        start_sec = int(start_str[0:2]) * 3600 + int(start_str[2:4]) * 60 + int(start_str[4:6])
        end_sec = int(end_str[0:2]) * 3600 + int(end_str[2:4]) * 60 + int(end_str[4:6])

        # Format date for display
        formatted_date = f"{date_str[0:4]}-{date_str[4:6]}-{date_str[6:8]}"

        return formatted_date, start_sec, end_sec

    return None, None, None


def parse_j40_video_times(filename: str) -> tuple:
    """
    Parse start time from J40 SD card video filename.
    Format: R1_YYYYMMDD_HHMMSS.mp4 (each file is ~5 minutes / 300 seconds)

    Returns (date_str, start_time_sec, end_time_sec) or (None, None, None) if parse fails.
    """
    match = re.match(r'(?:R[123]|Axis)_(\d{8})_(\d{6})\.mp4$', filename)
    if match:
        date_str = match.group(1)  # YYYYMMDD
        start_str = match.group(2)  # HHMMSS

        start_sec = int(start_str[0:2]) * 3600 + int(start_str[2:4]) * 60 + int(start_str[4:6])
        end_sec = start_sec + 300  # Each file is ~5 minutes

        formatted_date = f"{date_str[0:4]}-{date_str[4:6]}-{date_str[6:8]}"
        return formatted_date, start_sec, end_sec

    return None, None, None


@app.route('/api/recordings/dates')
def list_recording_dates():
    """List available recording dates across all cameras."""
    dates = set()

    if not RECORDINGS_DIR.exists():
        return jsonify({'error': 'Recordings drive not mounted', 'dates': []})

    # Check each camera folder (Reolink cameras: sd_R1, sd_R2, sd_R3)
    for camera_folder in ['sd_R1', 'sd_R2', 'sd_R3']:
        mp4_path = RECORDINGS_DIR / camera_folder / 'sdcard' / 'Mp4Record'
        if mp4_path.exists():
            for date_folder in mp4_path.iterdir():
                if date_folder.is_dir() and re.match(r'\d{4}-\d{2}-\d{2}', date_folder.name):
                    dates.add(date_folder.name)

    # Also check J40 SD card data (flat structure: R1/R1_YYYYMMDD_HHMMSS.mp4)
    if J40_SD_DATA_DIR.exists():
        for camera_folder in ['R1', 'R2', 'R3', 'Axis']:
            cam_path = J40_SD_DATA_DIR / camera_folder
            if cam_path.exists():
                for video_file in cam_path.glob('*.mp4'):
                    parsed_date, _, _ = parse_j40_video_times(video_file.name)
                    if parsed_date:
                        dates.add(parsed_date)

    return jsonify({'dates': sorted(dates, reverse=True)})


@app.route('/api/recordings/videos', methods=['POST'])
def list_recording_videos():
    """
    List videos from recordings drive for a specific camera and date.
    Optionally filter by time range based on Vola race timing.

    Request body:
    {
        "camera_id": "R1",
        "date": "2026-02-01",
        "start_time_sec": 36900,  // Optional: filter videos covering this time range
        "end_time_sec": 39600,    // Optional: filter videos covering this time range
    }
    """
    data = request.get_json() or {}
    camera_id = data.get('camera_id', 'R1')
    date = data.get('date')
    filter_start = data.get('start_time_sec')
    filter_end = data.get('end_time_sec')

    if not date:
        return jsonify({'error': 'date is required'}), 400

    # Map camera_id to folder name for Reolink recordings
    camera_folders = {
        'R1': 'sd_R1',
        'R2': 'sd_R2',
        'R3': 'sd_R3',
        'Axis': 'sd_Axis',
    }

    folder_name = camera_folders.get(camera_id)
    if not folder_name:
        return jsonify({'error': f'Unknown camera_id: {camera_id}'}), 400

    videos = []

    # Source 1: Reolink recordings drive (sd_R1/sdcard/Mp4Record/<date>/)
    videos_path = RECORDINGS_DIR / folder_name / 'sdcard' / 'Mp4Record' / date
    if videos_path.exists():
        for video_file in sorted(videos_path.glob('*.mp4')):
            parsed_date, start_sec, end_sec = parse_reolink_video_times(video_file.name)
            if start_sec is None:
                continue
            if filter_start is not None and filter_end is not None:
                if end_sec < filter_start or start_sec > filter_end:
                    continue
            start_time_str = f"{start_sec // 3600:02d}:{(start_sec % 3600) // 60:02d}:{start_sec % 60:02d}"
            end_time_str = f"{end_sec // 3600:02d}:{(end_sec % 3600) // 60:02d}:{end_sec % 60:02d}"
            videos.append({
                'name': video_file.name,
                'path': str(video_file),
                'start_time_sec': start_sec,
                'end_time_sec': end_sec,
                'start_time_str': start_time_str,
                'end_time_str': end_time_str,
                'duration_sec': end_sec - start_sec,
            })

    # Source 2: J40 SD card data (R1/R1_YYYYMMDD_HHMMSS.mp4)
    j40_cam_path = J40_SD_DATA_DIR / camera_id
    if j40_cam_path.exists():
        for video_file in sorted(j40_cam_path.glob('*.mp4')):
            parsed_date, start_sec, end_sec = parse_j40_video_times(video_file.name)
            if start_sec is None or parsed_date != date:
                continue
            if filter_start is not None and filter_end is not None:
                if end_sec < filter_start or start_sec > filter_end:
                    continue
            start_time_str = f"{start_sec // 3600:02d}:{(start_sec % 3600) // 60:02d}:{start_sec % 60:02d}"
            end_time_str = f"{end_sec // 3600:02d}:{(end_sec % 3600) // 60:02d}:{end_sec % 60:02d}"
            videos.append({
                'name': video_file.name,
                'path': str(video_file),
                'start_time_sec': start_sec,
                'end_time_sec': end_sec,
                'start_time_str': start_time_str,
                'end_time_str': end_time_str,
                'duration_sec': end_sec - start_sec,
            })

    if not videos:
        return jsonify({'error': f'No recordings found for {camera_id} on {date}', 'videos': []})

    # Sort all videos by start time
    videos.sort(key=lambda v: v['start_time_sec'])

    return jsonify({
        'camera_id': camera_id,
        'date': date,
        'videos': videos,
        'total': len(videos),
    })


@app.route('/api/recordings/videos-for-race', methods=['POST'])
def get_videos_for_race():
    """
    Get videos that cover a specific race's timing window.

    Request body:
    {
        "vola_file": "/path/to/vola.xlsx",
        "race": "U12 run 1",
        "camera_id": "R1",
        "num_athletes": 5,  // Number of athletes to process (0 or null = all)
        "startlist_file": "/path/to/start-list.pdf"  // Optional: PDF with racer names and teams
    }

    Returns list of videos needed and the racers they cover (with names/teams if startlist_file provided).
    """
    if not HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl not installed'}), 500

    data = request.get_json() or {}
    vola_file = data.get('vola_file')
    race = data.get('race', '').lower()
    camera_id = data.get('camera_id', 'R1')
    num_athletes = data.get('num_athletes', 0)  # 0 = all
    startlist_file = data.get('startlist_file')  # Optional PDF with names and teams

    # Parse start list PDF for names and teams if provided
    bib_to_racer = {}
    if startlist_file and Path(startlist_file).exists():
        bib_to_racer = parse_startlist_pdf(startlist_file)

    if not vola_file or not Path(vola_file).exists():
        return jsonify({'error': 'Invalid vola_file'}), 400

    # First, parse the Vola file to get racer timing data
    # Re-use the parse_vola_file logic
    race_map = {
        'u12 run 1': ('u12 start run 1', 'u12 end run 1'),
        'u12 run 2': ('u12 start run 2', 'u12 end run 2'),
        'u14 run 1': ('u14 start run 1', 'u14 end run 1'),
        'u14 run 2': ('u14 start run 2', 'u14 end run 2'),
    }

    if race not in race_map:
        return jsonify({'error': f'Invalid race: {race}'}), 400

    start_sheet, end_sheet = race_map[race]
    camera_offset_pct = CAMERAS.get(camera_id, {}).get('offset_pct', 0) / 100.0

    try:
        wb = openpyxl.load_workbook(vola_file, data_only=True)

        # Parse start times
        start_times = {}
        ws_start = wb[start_sheet]
        for row in ws_start.iter_rows(min_row=2, values_only=True):
            bib_raw = row[1]
            time_str = row[2]
            if bib_raw and time_str:
                try:
                    bib = int(bib_raw)
                except (ValueError, TypeError):
                    continue
                start_seconds = parse_vola_time(str(time_str))
                if start_seconds is not None:
                    start_times[bib] = start_seconds

        # Parse end times
        end_times = {}
        run_durations = {}
        ws_end = wb[end_sheet]
        for row in ws_end.iter_rows(min_row=2, values_only=True):
            bib_raw = row[1]
            time_str = row[2]
            duration = row[3] if len(row) > 3 else None
            if bib_raw:
                try:
                    bib = int(bib_raw)
                except (ValueError, TypeError):
                    continue
                if time_str:
                    end_seconds = parse_vola_time(str(time_str))
                    if end_seconds is not None:
                        end_times[bib] = end_seconds
                if duration is not None and isinstance(duration, (int, float)):
                    run_durations[bib] = float(duration)

        wb.close()

        # Build racer list with camera timing windows
        racers = []
        for bib in sorted(start_times.keys()):
            start_sec = start_times[bib]

            run_duration = None
            if bib in end_times:
                run_duration = end_times[bib] - start_sec
            elif bib in run_durations:
                run_duration = run_durations[bib]

            if run_duration and run_duration > 0:
                camera_start_sec = start_sec + (run_duration * camera_offset_pct)
                camera_duration = run_duration * 0.2
                camera_end_sec = camera_start_sec + camera_duration
            else:
                estimated_duration = 40.0
                camera_start_sec = start_sec + (estimated_duration * camera_offset_pct)
                camera_duration = estimated_duration * 0.2
                camera_end_sec = camera_start_sec + camera_duration

            # Get name, team, and gender from start list
            racer_info = bib_to_racer.get(bib, {})
            racers.append({
                'bib': bib,
                'name': racer_info.get('name', ''),  # Name from start list PDF
                'team': racer_info.get('team', ''),  # Team from start list PDF
                'gender': racer_info.get('gender', ''),  # Gender from start list PDF
                'start_time_sec': start_sec,
                'run_duration': run_duration,  # Actual run time from Vola (for start offset calculation)
                'camera_start_sec': camera_start_sec,
                'camera_end_sec': camera_end_sec,
                'finished': bib in end_times,
            })

        # Sort by start time
        racers.sort(key=lambda r: r['start_time_sec'])

        # Limit to num_athletes if specified
        if num_athletes and num_athletes > 0:
            racers = racers[:num_athletes]

        if not racers:
            return jsonify({'error': 'No racers found', 'videos': [], 'racers': []})

        # Find the time range we need videos for
        first_camera_time = min(r['camera_start_sec'] for r in racers)
        last_camera_time = max(r['camera_end_sec'] for r in racers)

        # Add some buffer (30 seconds before and after)
        filter_start = first_camera_time - 30
        filter_end = last_camera_time + 30

        # Extract date from Vola filename (format: Vola_export_U12_U14_02-01-2026.xlsx)
        vola_name = Path(vola_file).stem
        date_match = re.search(r'(\d{2})-(\d{2})-(\d{4})', vola_name)
        if date_match:
            race_date = f"{date_match.group(3)}-{date_match.group(1)}-{date_match.group(2)}"
        else:
            return jsonify({'error': 'Could not extract date from Vola filename'}), 400

        # Get videos that cover this time range
        videos_response = list_recording_videos()
        # Actually call the function with proper request context
        from flask import request as flask_request

        # Build path to videos
        camera_folders = {'R1': 'sd_R1', 'R2': 'sd_R2', 'R3': 'sd_R3', 'Axis': 'sd_Axis'}
        folder_name = camera_folders.get(camera_id, 'sd_R1')
        videos_path = RECORDINGS_DIR / folder_name / 'sdcard' / 'Mp4Record' / race_date

        videos = []
        if videos_path.exists():
            for video_file in sorted(videos_path.glob('*.mp4')):
                parsed_date, start_sec, end_sec = parse_reolink_video_times(video_file.name)
                if start_sec is None:
                    continue

                # Check if video overlaps with the time range we need
                if end_sec < filter_start or start_sec > filter_end:
                    continue

                start_time_str = f"{start_sec // 3600:02d}:{(start_sec % 3600) // 60:02d}:{start_sec % 60:02d}"
                end_time_str = f"{end_sec // 3600:02d}:{(end_sec % 3600) // 60:02d}:{end_sec % 60:02d}"

                videos.append({
                    'name': video_file.name,
                    'path': str(video_file),
                    'start_time_sec': start_sec,
                    'end_time_sec': end_sec,
                    'start_time_str': start_time_str,
                    'end_time_str': end_time_str,
                })

        return jsonify({
            'race': race,
            'camera_id': camera_id,
            'date': race_date,
            'num_athletes': len(racers),
            'time_range': {
                'start_sec': filter_start,
                'end_sec': filter_end,
                'start_str': f"{int(filter_start) // 3600:02d}:{(int(filter_start) % 3600) // 60:02d}:{int(filter_start) % 60:02d}",
                'end_str': f"{int(filter_end) // 3600:02d}:{(int(filter_end) % 3600) // 60:02d}:{int(filter_end) % 60:02d}",
            },
            'videos': videos,
            'racers': racers,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/process/video', methods=['POST'])
def process_video():
    """Start processing video file(s) - returns job_id for tracking."""
    data = request.json
    config_path = data.get('config_path')
    video_path = data.get('video_path')  # Optional - can use videos from config instead
    use_config_videos = data.get('use_config_videos', False)

    print(f"[DEBUG] process_video called: config_path={config_path}, video_path={video_path}, use_config_videos={use_config_videos}")

    if not config_path:
        return jsonify({'error': 'Missing config_path'}), 400

    if not Path(config_path).exists():
        return jsonify({'error': f'Config not found: {config_path}'}), 404

    # If not using config videos, need a video_path
    if not use_config_videos and not video_path:
        return jsonify({'error': 'Missing video_path or use_config_videos'}), 400

    if video_path and not Path(video_path).exists():
        return jsonify({'error': f'Video not found: {video_path}'}), 404

    # Generate job ID
    job_id = str(uuid.uuid4())[:8]

    # Run the runner.py script as a background process
    runner_path = Path(__file__).resolve().parent / 'runner.py'

    try:
        if use_config_videos:
            # Use --use-config-videos flag to process videos from config's vola_videos list
            cmd = [PYTHON, '-u', str(runner_path), '--config', config_path, '--use-config-videos']
        else:
            cmd = [PYTHON, '-u', str(runner_path), '--config', config_path, video_path]
        print(f"[DEBUG] Running command: {' '.join(cmd)}")

        process = Popen(
            cmd,
            stdout=PIPE,
            stderr=PIPE,
            text=True,
            bufsize=1  # Line buffered
        )

        with jobs_lock:
            active_jobs[job_id] = {
                'process': process,
                'status': 'running',
                'output': '',
                'config_path': config_path,
                'video_path': video_path,
                'started_at': datetime.now().isoformat(),
            }

        # Start a thread to collect output
        def collect_output():
            try:
                for line in process.stdout:
                    print(f"[DEBUG] stdout: {line.rstrip()}")
                    with jobs_lock:
                        if job_id in active_jobs:
                            active_jobs[job_id]['output'] += line
                process.wait()
                # Collect any stderr
                stderr_output = process.stderr.read() if process.stderr else ''
                if stderr_output:
                    print(f"[DEBUG] stderr: {stderr_output}")
                with jobs_lock:
                    if job_id in active_jobs:
                        if process.returncode == 0:
                            active_jobs[job_id]['status'] = 'completed'
                        elif process.returncode == -signal.SIGTERM or process.returncode == -signal.SIGKILL:
                            active_jobs[job_id]['status'] = 'stopped'
                        else:
                            active_jobs[job_id]['status'] = 'failed'
                            active_jobs[job_id]['error'] = stderr_output or f'Exit code: {process.returncode}'
                        print(f"[DEBUG] Job {job_id} finished with status: {active_jobs[job_id]['status']}")
            except Exception as e:
                print(f"[DEBUG] Exception in collect_output: {e}")
                with jobs_lock:
                    if job_id in active_jobs:
                        active_jobs[job_id]['status'] = 'failed'
                        active_jobs[job_id]['error'] = str(e)

        thread = threading.Thread(target=collect_output, daemon=True)
        thread.start()

        return jsonify({
            'job_id': job_id,
            'status': 'running',
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/process/rtsp', methods=['POST'])
def process_rtsp():
    """Start live RTSP monitoring using shared stream manager.

    Multiple sessions on the same camera share a single RTSP connection.
    Frames are decoded once and distributed to all detection engines.
    """
    data = request.json
    config_path = data.get('config_path')

    if not config_path:
        return jsonify({'error': 'Missing config_path'}), 400

    if not Path(config_path).exists():
        return jsonify({'error': f'Config not found: {config_path}'}), 404

    # Load config to get camera_id and session_id
    try:
        with open(config_path) as f:
            config = json.load(f)
    except Exception as e:
        return jsonify({'error': f'Failed to read config: {e}'}), 500

    camera_id = config.get('camera_id', '')
    session_id = config.get('session_id', Path(config_path).stem)
    rtsp_url = config.get('camera_url', '')

    if not rtsp_url:
        # Fall back to CAMERAS dict
        rtsp_url = CAMERAS.get(camera_id, {}).get('rtsp_url', '')
    if not rtsp_url:
        return jsonify({'error': f'No RTSP URL for camera {camera_id}'}), 400

    # Generate job ID
    job_id = str(uuid.uuid4())[:8]

    output_dir = str(Path(__file__).resolve().parent.parent / 'output')

    try:
        # Create runner in-process (no subprocess needed)
        runner = SkiFramesRunner(config_path, output_dir)

        # Register with stream manager (shares RTSP connection per camera)
        stream_manager.start_session(camera_id, rtsp_url, session_id, runner)

        # Track session
        with rtsp_sessions_lock:
            rtsp_sessions[job_id] = {
                'runner': runner,
                'session_id': session_id,
                'camera_id': camera_id,
                'config_path': config_path,
                'status': 'running',
                'started_at': datetime.now().isoformat(),
            }

        print(f"[RTSP] Session {session_id} started on camera {camera_id} "
              f"(job {job_id})")

        return jsonify({
            'job_id': job_id,
            'status': 'running',
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/process/status/<job_id>')
def process_status(job_id):
    """Get the status of a processing job (subprocess or in-process RTSP)."""
    # Check in-process RTSP sessions first
    with rtsp_sessions_lock:
        rtsp_session = rtsp_sessions.get(job_id)

    if rtsp_session:
        runner = rtsp_session['runner']
        session_id = rtsp_session['session_id']
        camera_id = rtsp_session['camera_id']

        # Check if session is still active in stream manager
        runs_detected = runner.engine.run_count if hasattr(runner.engine, 'run_count') else 0
        actual_status = rtsp_session['status']

        # Check if stream is still running for this session
        if actual_status == 'running':
            sm_camera = stream_manager.get_session_camera(session_id)
            if sm_camera is None:
                # Session was removed (expired or stopped)
                actual_status = 'completed'
                rtsp_session['status'] = actual_status

        response = {
            'job_id': job_id,
            'status': actual_status,
            'runs_detected': runs_detected,
            'output_dir': runner.session_dir if hasattr(runner, 'session_dir') else '',
            'output': f'[In-process RTSP session on {camera_id}]\nRuns detected: {runs_detected}',
            'error': rtsp_session.get('error', ''),
        }
        return jsonify(response)

    # Fall back to subprocess jobs
    with jobs_lock:
        if job_id not in active_jobs:
            return jsonify({'error': 'Job not found'}), 404

        job = active_jobs[job_id]
        output = job['output']

        # Parse output for progress info
        runs_detected = 0
        output_dir = ''
        for line in output.split('\n'):
            # Count completed runs in real-time from "[RUN X] Completed" messages
            if '] Completed at' in line:
                runs_detected += 1
            # Also check final summary line "Runs detected: X"
            if 'Runs detected:' in line:
                try:
                    runs_detected = int(line.split(':')[1].strip())
                except:
                    pass
            if 'Output directory:' in line:
                output_dir = line.split(':', 1)[1].strip()

        # Health check: verify subprocess is actually alive if status says 'running'
        actual_status = job['status']
        if actual_status == 'running' and 'process' in job:
            poll = job['process'].poll()
            if poll is not None:
                # Process has exited but status wasn't updated
                if poll == 0:
                    actual_status = 'completed'
                elif poll == -15 or poll == -9:  # SIGTERM or SIGKILL
                    actual_status = 'stopped'
                else:
                    actual_status = 'failed'
                    job['error'] = f'Process exited with code {poll}'
                job['status'] = actual_status
                print(f"[HEALTH] Job {job_id} process died (exit={poll}), status corrected to '{actual_status}'")

        response = {
            'job_id': job_id,
            'status': actual_status,
            'runs_detected': runs_detected,
            'output_dir': output_dir,
            'output': output,
            'error': job.get('error', ''),
        }
        print(f"[DEBUG] Status for {job_id}: status={actual_status}, runs={runs_detected}")
        return jsonify(response)


@app.route('/api/process/stop/<job_id>', methods=['POST'])
def stop_process(job_id):
    """Stop a running processing job (subprocess or in-process RTSP)."""
    # Check in-process RTSP sessions first
    with rtsp_sessions_lock:
        rtsp_session = rtsp_sessions.get(job_id)

    if rtsp_session:
        session_id = rtsp_session['session_id']
        stream_manager.stop_session(session_id)
        with rtsp_sessions_lock:
            rtsp_sessions[job_id]['status'] = 'stopped'
        print(f"[STOP] In-process RTSP session {session_id} stopped (job {job_id})")
        return jsonify({'success': True, 'status': 'stopped'})

    # Fall back to subprocess jobs
    with jobs_lock:
        if job_id not in active_jobs:
            return jsonify({'error': 'Job not found'}), 404

        job = active_jobs[job_id]
        if job['status'] != 'running':
            return jsonify({'error': f"Job is not running (status: {job['status']})"}), 400

        process = job['process']

    # Send SIGTERM to gracefully stop
    try:
        process.terminate()
        # Give it a moment to terminate gracefully
        try:
            process.wait(timeout=3)
        except:
            # Force kill if it doesn't stop
            process.kill()

        with jobs_lock:
            if job_id in active_jobs:
                active_jobs[job_id]['status'] = 'stopped'

        return jsonify({'success': True, 'status': 'stopped'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/streams/status')
def streams_status():
    """Get status of shared RTSP streams and their sessions."""
    return jsonify(stream_manager.get_status())


# =============================================================================
# MONTAGE GALLERY ENDPOINTS
# =============================================================================

@app.route('/api/montages/latest')
def montages_latest():
    """List recent full-res montage images across all sessions, newest first."""
    output_dir = Path(__file__).resolve().parent.parent / 'output'
    if not output_dir.exists():
        return jsonify({'montages': []})

    montages = []
    # Walk all session directories looking for fullres images
    for session_dir in output_dir.iterdir():
        if not session_dir.is_dir():
            continue
        session_id = session_dir.name
        # Find all fullres jpg files recursively
        for fullres_file in session_dir.rglob('fullres/*.jpg'):
            # Skip thumbnails
            if '_thumb' in fullres_file.name:
                continue
            stat = fullres_file.stat()
            # Build relative path from output dir for serving
            rel_path = fullres_file.relative_to(output_dir)
            # Extract FPS from filename (e.g., "run_001_093523_4.0fps.jpg" -> 4.0)
            import re as _re
            fps_match = _re.search(r'_(\d+\.?\d*)fps', fullres_file.name)
            fps_val = float(fps_match.group(1)) if fps_match else None

            montages.append({
                'filename': fullres_file.name,
                'path': str(rel_path),
                'session_id': session_id,
                'size_kb': int(stat.st_size / 1024),
                'modified': stat.st_mtime,
                'fps': fps_val,
            })

    # Sort by modification time, newest first
    montages.sort(key=lambda m: m['modified'], reverse=True)

    # Limit to most recent 10
    montages = montages[:10]

    return jsonify({'montages': montages})


@app.route('/api/montages/image/<path:image_path>')
def montages_image(image_path):
    """Serve a montage image from the output directory."""
    output_dir = Path(__file__).resolve().parent.parent / 'output'
    full_path = output_dir / image_path

    # Security: ensure path doesn't escape output directory
    try:
        full_path.resolve().relative_to(output_dir.resolve())
    except ValueError:
        return jsonify({'error': 'Invalid path'}), 403

    if not full_path.exists():
        return jsonify({'error': 'Image not found'}), 404

    return send_file(str(full_path), mimetype='image/jpeg')


# =============================================================================
# VIDEO STITCH ENDPOINTS
# =============================================================================

# Track active stitch jobs
stitch_jobs = {}
stitch_jobs_lock = threading.Lock()


@app.route('/api/stitch/logos')
def list_available_logos():
    """
    List all available logo images for overlay selection.

    Returns list of logos with filename, name, and preview URL.
    User can select which logos to include and in what order (left to right).
    """
    logos = []
    if LOGOS_DIR.exists():
        for f in sorted(LOGOS_DIR.glob('*.png')):
            # Skip hidden files
            if f.name.startswith('.'):
                continue

            # Create a friendly display name from filename
            # e.g., "US-Ski-Snowboard.png" -> "US Ski Snowboard"
            display_name = f.stem.replace('_', ' ').replace('-', ' ')

            logos.append({
                'filename': f.name,
                'name': display_name,
                'path': str(f),
                'preview_url': f'/api/stitch/logo/{f.name}'
            })

    return jsonify({
        'logos': logos,
        'default_order': [
            'US-Ski-Snowboard.png',
            'NHARA_logo.png',
            'RMST_logo.png',
            'Ragged_logo.png',
            'Skiframes-com_logo.png'
        ],
        'excluded_by_default': [
            'skieast_logo.png'
        ]
    })


@app.route('/api/stitch/logo/<filename>')
def get_logo_preview(filename):
    """Serve a logo image for preview."""
    # Sanitize filename to prevent path traversal
    safe_filename = Path(filename).name
    logo_path = LOGOS_DIR / safe_filename

    if not logo_path.exists() or not logo_path.suffix.lower() == '.png':
        return jsonify({'error': 'Logo not found'}), 404

    return send_file(logo_path, mimetype='image/png')


@app.route('/api/stitch/configs')
def list_stitch_configs():
    """List saved stitch cut point configurations."""
    configs = []
    for f in STITCH_CONFIG_DIR.glob('*.json'):
        try:
            with open(f, 'r') as fp:
                config = json.load(fp)
                configs.append({
                    'filename': f.name,
                    'name': config.get('name', f.stem),
                    'created': config.get('created', ''),
                    'cuts': config.get('cuts', [])
                })
        except Exception as e:
            print(f"Error reading stitch config {f}: {e}")
    return jsonify(sorted(configs, key=lambda x: x['name']))


@app.route('/api/stitch/config/<name>')
def get_stitch_config(name):
    """Get a specific stitch configuration."""
    # Sanitize name to prevent path traversal
    safe_name = re.sub(r'[^\w\-]', '_', name.lower())
    config_path = STITCH_CONFIG_DIR / f"{safe_name}.json"

    if not config_path.exists():
        return jsonify({'error': 'Config not found'}), 404

    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
        return jsonify(config)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/stitch/config', methods=['POST'])
def save_stitch_config():
    """
    Save a stitch cut point configuration.

    Request body:
    {
        "name": "U12 run 1",
        "cuts": [
            {"camera": "R1", "start_pct": -0.05, "end_pct": 0.05},
            {"camera": "Axis", "start_pct": 0.05, "end_pct": 0.425},
            {"camera": "R2", "start_pct": 0.425, "end_pct": 0.575},
            {"camera": "R3", "start_pct": 0.575, "end_pct": 1.05}
        ]
    }
    """
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    cuts = data.get('cuts', [])

    if not name:
        return jsonify({'error': 'Name is required'}), 400

    if not cuts or len(cuts) == 0:
        return jsonify({'error': 'At least one cut is required'}), 400

    # Validate cuts
    for cut in cuts:
        if 'camera' not in cut or 'start_pct' not in cut or 'end_pct' not in cut:
            return jsonify({'error': 'Each cut must have camera, start_pct, and end_pct'}), 400

    # Create config
    config = {
        'name': name,
        'created': datetime.now().isoformat(),
        'cuts': cuts
    }

    # Save to file
    safe_name = re.sub(r'[^\w\-]', '_', name.lower())
    config_path = STITCH_CONFIG_DIR / f"{safe_name}.json"

    try:
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=2)
        return jsonify({'success': True, 'path': str(config_path)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/stitch/parse-racers', methods=['POST'])
def parse_racers_for_stitch():
    """
    Parse Vola file, start list PDF, and results PDF to get racer data for stitching.

    Request body:
    {
        "vola_file": "/path/to/vola.xlsx",
        "race": "U12 run 1",
        "startlist_file": "/path/to/startlist.pdf",  // Optional
        "results_file": "/path/to/results.pdf",  // Optional - for USSA IDs, rankings, DSQ/DNF
        "include_dsq_dnf": true  // Optional - include DSQ/DNF racers with 60s replacement time
    }

    Returns list of racers with timing, name, USSA ID, ranking, and status data.
    """
    if not HAS_OPENPYXL:
        return jsonify({'error': 'openpyxl not installed'}), 500

    data = request.get_json() or {}
    vola_file = data.get('vola_file')
    race = data.get('race', '').lower()
    startlist_file = data.get('startlist_file')
    results_file = data.get('results_file')
    include_dsq_dnf = data.get('include_dsq_dnf', True)

    if not vola_file or not Path(vola_file).exists():
        return jsonify({'error': 'Invalid vola_file'}), 400

    # Map race selection to sheet names
    race_map = {
        'u12 run 1': ('u12 start run 1', 'u12 end run 1'),
        'u12 run 2': ('u12 start run 2', 'u12 end run 2'),
        'u14 run 1': ('u14 start run 1', 'u14 end run 1'),
        'u14 run 2': ('u14 start run 2', 'u14 end run 2'),
    }

    if race not in race_map:
        return jsonify({'error': f'Invalid race. Must be one of: {list(race_map.keys())}'}), 400

    # Parse start list for names if provided
    bib_to_startlist = {}
    if startlist_file and Path(startlist_file).exists():
        bib_to_startlist = parse_startlist_pdf(startlist_file)

    # Parse results PDF for USSA IDs, rankings, and status
    bib_to_results = {}
    if results_file and Path(results_file).exists():
        bib_to_results = parse_results_pdf(results_file)

    start_sheet, end_sheet = race_map[race]

    # Default replacement time for DSQ/DNF (60 seconds)
    DSQ_DNF_REPLACEMENT_TIME = 60.0

    try:
        wb = openpyxl.load_workbook(vola_file, data_only=True)

        # Parse start times
        start_times = {}
        ws_start = wb[start_sheet]
        for row in ws_start.iter_rows(min_row=2, values_only=True):
            bib_raw = row[1]
            time_str = row[2]
            if bib_raw and time_str:
                try:
                    bib = int(bib_raw)
                except (ValueError, TypeError):
                    continue
                start_seconds = parse_vola_time(str(time_str))
                if start_seconds is not None:
                    start_times[bib] = start_seconds

        # Parse end times and durations
        end_times = {}
        run_durations = {}
        ws_end = wb[end_sheet]
        for row in ws_end.iter_rows(min_row=2, values_only=True):
            bib_raw = row[1]
            time_str = row[2]
            duration = row[3] if len(row) > 3 else None
            if bib_raw:
                try:
                    bib = int(bib_raw)
                except (ValueError, TypeError):
                    continue
                if time_str:
                    end_seconds = parse_vola_time(str(time_str))
                    if end_seconds is not None:
                        end_times[bib] = end_seconds
                if duration is not None and isinstance(duration, (int, float)):
                    run_durations[bib] = float(duration)

        wb.close()

        # Build racer list
        racers = []
        for bib in sorted(start_times.keys()):
            start_sec = start_times[bib]

            # Get data from results PDF (has USSA ID, rank, status)
            results_info = bib_to_results.get(bib, {})
            startlist_info = bib_to_startlist.get(bib, {})

            # Determine status from results
            status = results_info.get('status', 'finished')
            rank = results_info.get('rank')

            # Calculate run duration
            run_duration = None
            if bib in end_times:
                run_duration = end_times[bib] - start_sec
            elif bib in run_durations:
                run_duration = run_durations[bib]

            # Handle DSQ/DNF/DNS - use replacement time
            if run_duration is None or run_duration <= 0:
                if status in ('DSQ', 'DNF', 'DNS') and include_dsq_dnf:
                    run_duration = DSQ_DNF_REPLACEMENT_TIME
                else:
                    continue  # Skip racers without valid finish and not DSQ/DNF

            # Get name/team - prefer results (has parsed name), fallback to startlist
            name = results_info.get('name') or startlist_info.get('name', f'Bib{bib}')
            team = results_info.get('team') or startlist_info.get('team', '')
            gender = results_info.get('gender') or startlist_info.get('gender', '') or get_gender_from_bib(bib)

            # Get USSA ID from results
            ussa_id = results_info.get('ussa_id', '')

            racer = {
                'bib': bib,
                'name': name,
                'team': team,
                'gender': gender,
                'ussa_id': ussa_id,
                'ussa_profile_url': f"https://www.usskiandsnowboard.org/public-tools/members/{ussa_id.lstrip('E')}" if ussa_id else None,
                'start_time_sec': start_sec,
                'start_time_str': seconds_to_time_str(start_sec),
                'finish_time_sec': start_sec + run_duration,
                'finish_time_str': seconds_to_time_str(start_sec + run_duration),
                'run_duration': run_duration,
                'status': status,  # 'finished', 'DSQ', 'DNF', 'DNS'
                'rank': rank,  # None for DSQ/DNF/DNS
            }
            racers.append(racer)

        # Sort by start time
        racers.sort(key=lambda r: r['start_time_sec'])

        # Find videos for all 4 cameras
        # Extract date from Vola filename (format: Vola_export_U12_U14_02-01-2026.xlsx)
        vola_name = Path(vola_file).stem
        date_match = re.search(r'(\d{2})-(\d{2})-(\d{4})', vola_name)
        if not date_match:
            return jsonify({'error': 'Could not extract date from Vola filename'}), 400

        race_date = f"{date_match.group(3)}-{date_match.group(1)}-{date_match.group(2)}"

        # Calculate time range needed (earliest start - 60s buffer, latest finish + 60s buffer)
        if racers:
            earliest_start = min(r['start_time_sec'] for r in racers) - 60
            latest_finish = max(r['finish_time_sec'] for r in racers) + 60
        else:
            earliest_start = 0
            latest_finish = 86400

        # Find videos for each camera
        camera_folders = {'R1': 'sd_R1', 'R2': 'sd_R2', 'R3': 'sd_R3', 'Axis': 'sd_Axis'}
        video_paths = {}

        for camera_id, folder_name in camera_folders.items():
            camera_videos = []

            if camera_id == 'Axis':
                # Axis has different folder structure: sd_Axis/sdcard/YYYYMMDD/HH/YYYYMMDD_HHMMSS_XXXX_XXX/YYYYMMDD_HH/*.mkv
                axis_date_str = race_date.replace('-', '')  # 2026-02-01 -> 20260201
                axis_base = RECORDINGS_DIR / folder_name / 'sdcard' / axis_date_str
                if axis_base.exists():
                    # Search through hour folders
                    for hour_folder in sorted(axis_base.glob('*')):
                        if not hour_folder.is_dir():
                            continue
                        for recording_folder in sorted(hour_folder.glob('*')):
                            if not recording_folder.is_dir():
                                continue
                            # Find mkv files in subdirectories
                            for mkv_file in recording_folder.glob('**/*.mkv'):
                                # Parse Axis filename: YYYYMMDD_HHMMSS_XXXX.mkv
                                fname = mkv_file.stem  # e.g., 20260201_081402_A4CC
                                match = re.match(r'(\d{8})_(\d{6})_', fname)
                                if match:
                                    start_str = match.group(2)
                                    start_sec = int(start_str[0:2]) * 3600 + int(start_str[2:4]) * 60 + int(start_str[4:6])
                                    # Estimate end time (Axis videos are typically ~5 min segments)
                                    end_sec = start_sec + 300

                                    if end_sec >= earliest_start and start_sec <= latest_finish:
                                        camera_videos.append({
                                            'path': str(mkv_file),
                                            'start_sec': start_sec,
                                            'end_sec': end_sec,
                                        })
            else:
                # Reolink cameras: sd_XX/sdcard/Mp4Record/YYYY-MM-DD/*.mp4
                videos_path = RECORDINGS_DIR / folder_name / 'sdcard' / 'Mp4Record' / race_date
                if videos_path.exists():
                    for video_file in sorted(videos_path.glob('*.mp4')):
                        parsed_date, start_sec, end_sec = parse_reolink_video_times(video_file.name)
                        if start_sec is None:
                            continue

                        if end_sec >= earliest_start and start_sec <= latest_finish:
                            camera_videos.append({
                                'path': str(video_file),
                                'start_sec': start_sec,
                                'end_sec': end_sec,
                            })

            # Return ALL videos that overlap with the race time range, sorted by start time
            if camera_videos:
                camera_videos.sort(key=lambda v: v['start_sec'])
                video_paths[camera_id] = [v['path'] for v in camera_videos]

        return jsonify({
            'race': race,
            'race_date': race_date,
            'racers': racers,
            'total_racers': len(racers),
            'video_paths': video_paths,  # Now a dict of camera_id -> list of video paths
            'time_range': {
                'start_sec': earliest_start,
                'end_sec': latest_finish,
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def run_stitch_job(job_id: str, config: dict):
    """Background thread function to run the stitch job."""
    from stitcher import VideoStitcher, CameraCut, Racer

    try:
        with stitch_jobs_lock:
            stitch_jobs[job_id]['status'] = 'running'
            stitch_jobs[job_id]['progress'] = {'current': 0, 'total': 0, 'current_racer': ''}

        # Extract config
        racers_data = config['racers']
        cuts_data = config['cuts']
        video_paths = config['video_paths']
        output_dir = config.get('output_dir', str(Path(__file__).parent.parent / 'output'))
        race_name = config.get('race_name', 'race')
        race_title = config.get('race_title', '')
        race_info = config.get('race_info', {})
        selected_logos = config.get('selected_logos')
        generate_comparison = config.get('generate_comparison', False)

        # Convert to objects
        cuts = [CameraCut(**c) for c in cuts_data]
        racers = []
        for r in racers_data:
            if r.get('run_duration', 0) <= 0:
                continue
            racers.append(Racer(
                bib=r['bib'],
                name=r.get('name', f"Bib{r['bib']}"),
                team=r.get('team', ''),
                gender=r.get('gender', ''),
                start_time_sec=r['start_time_sec'],
                finish_time_sec=r.get('finish_time_sec', r['start_time_sec'] + r['run_duration']),
                duration=r['run_duration'],
                ussa_id=r.get('ussa_id', ''),
                status=r.get('status', 'finished')
            ))

        with stitch_jobs_lock:
            stitch_jobs[job_id]['progress']['total'] = len(racers)

        # Stop flag function - checks if job should stop
        def should_stop():
            with stitch_jobs_lock:
                return stitch_jobs.get(job_id, {}).get('status') == 'stopping'

        # Create stitcher and run
        stitcher = VideoStitcher(
            racers=racers,
            cuts=cuts,
            video_paths=video_paths,
            output_dir=output_dir,
            race_name=race_name,
            race_title=race_title,
            race_info=race_info,
            selected_logos=selected_logos,
            stop_flag=should_stop
        )

        def progress_callback(current, total, name):
            with stitch_jobs_lock:
                stitch_jobs[job_id]['progress'] = {
                    'current': current,
                    'total': total,
                    'current_racer': name
                }

        # Get parallel processing options
        # Default workers: 8 for M1 Max, override via env SKIFRAMES_MAX_WORKERS for server
        default_workers = int(os.environ.get('SKIFRAMES_MAX_WORKERS', 8))
        parallel = config.get('parallel', True)
        max_workers = config.get('max_workers', default_workers)

        if parallel:
            print(f"Using parallel processing with {max_workers} workers")
            if generate_comparison:
                outputs = stitcher.process_all_with_comparison_parallel(
                    max_workers=max_workers,
                    progress_callback=progress_callback,
                    generate_comparison=True
                )
            else:
                outputs = stitcher.process_all_parallel(
                    max_workers=max_workers,
                    progress_callback=progress_callback
                )
        else:
            print("Using sequential processing")
            if generate_comparison:
                outputs = stitcher.process_all_with_comparison(progress_callback=progress_callback, generate_comparison=True)
            else:
                outputs = stitcher.process_all(progress_callback=progress_callback)

        with stitch_jobs_lock:
            # Check if we were stopped
            if stitch_jobs[job_id]['status'] == 'stopping':
                stitch_jobs[job_id]['status'] = 'stopped'
            else:
                stitch_jobs[job_id]['status'] = 'completed'
            stitch_jobs[job_id]['outputs'] = [str(p) for p in outputs]
            stitch_jobs[job_id]['output_dir'] = output_dir

    except Exception as e:
        with stitch_jobs_lock:
            stitch_jobs[job_id]['status'] = 'error'
            stitch_jobs[job_id]['error'] = str(e)


@app.route('/api/stitch/process', methods=['POST'])
def start_stitch_process():
    """
    Start video stitching process.

    Request body:
    {
        "racers": [...],  // From parse-racers endpoint
        "cuts": [...],    // Cut point configuration
        "video_paths": {  // Camera ID to video file path
            "R1": "/path/to/r1.mp4",
            "R2": "/path/to/r2.mp4",
            "Axis": "/path/to/axis.mp4",
            "R3": "/path/to/r3.mp4"
        },
        "race_name": "u12_run1",  // Optional
        "race_title": "Western Division Ranking - SL",  // Optional, for video overlay (deprecated, use race_info)
        "race_info": {  // Optional, for manifest metadata and title overlay
            "event": "Western Division Ranking",
            "discipline": "SL",
            "age_group": "U14",
            "run": "Run 1",
            "date": "Sunday, 2026/02/01",
            "course": "Flying Yankee",
            "location": "Ragged Mountain, NH",
            "type": "USSA/NHARA",
            "vertical_drop": "85m",
            "length": "305m",
            "gates": "35",
            "snow": "Packed Powder"
        },
        "selected_logos": [  // Optional, logo filenames in order (left to right)
            "US-Ski-Snowboard.png",
            "NHARA_logo.png",
            "RMST_logo.png",
            "Ragged_logo.png",
            "Skiframes-com_logo.png"
        ],
        "output_dir": "/path/to/output",  // Optional
        "generate_comparison": false,  // Optional, generate vs fastest videos
        "parallel": true,  // Optional, use parallel processing (default: true for M1 Mac)
        "max_workers": 4   // Optional, number of concurrent workers (default: 4 for M1 Max)
    }
    """
    data = request.get_json() or {}

    racers = data.get('racers', [])
    cuts = data.get('cuts', [])
    video_paths = data.get('video_paths', {})

    if not racers:
        return jsonify({'error': 'No racers provided'}), 400
    if not cuts:
        return jsonify({'error': 'No cuts provided'}), 400
    if not video_paths:
        return jsonify({'error': 'No video_paths provided'}), 400

    # Validate video paths exist (video_paths values can be single path or list of paths)
    for camera, paths in video_paths.items():
        # Normalize to list
        if isinstance(paths, str):
            paths = [paths]
        if not paths:
            return jsonify({'error': f'No video files provided for {camera}'}), 400
        for path in paths:
            if not Path(path).exists():
                return jsonify({'error': f'Video file not found for {camera}: {path}'}), 400

    # Create job
    job_id = str(uuid.uuid4())[:8]

    with stitch_jobs_lock:
        stitch_jobs[job_id] = {
            'status': 'starting',
            'progress': {'current': 0, 'total': len(racers), 'current_racer': ''},
            'outputs': [],
            'error': None
        }

    # Start background thread
    config = {
        'racers': racers,
        'cuts': cuts,
        'video_paths': video_paths,
        'race_name': data.get('race_name', 'race'),
        'race_title': data.get('race_title', ''),  # e.g., "Western Division U14 Ranking - SL"
        'race_info': data.get('race_info', {}),  # Structured race metadata for manifest
        'selected_logos': data.get('selected_logos'),  # Logo filenames in order (left to right)
        'output_dir': data.get('output_dir', str(Path(__file__).parent.parent / 'output')),
        'generate_comparison': data.get('generate_comparison', False)  # Generate videos vs fastest racer
    }

    thread = threading.Thread(target=run_stitch_job, args=(job_id, config))
    thread.daemon = True
    thread.start()

    return jsonify({'job_id': job_id, 'status': 'starting'})


@app.route('/api/stitch/job/<job_id>')
def get_stitch_job_status(job_id):
    """Get status of a stitch job."""
    with stitch_jobs_lock:
        if job_id not in stitch_jobs:
            return jsonify({'error': 'Job not found'}), 404

        job = stitch_jobs[job_id].copy()

    return jsonify(job)


@app.route('/api/stitch/job/<job_id>/stop', methods=['POST'])
def stop_stitch_job(job_id):
    """Stop a running stitch job (marks as stopped, thread will check)."""
    with stitch_jobs_lock:
        if job_id not in stitch_jobs:
            return jsonify({'error': 'Job not found'}), 404

        stitch_jobs[job_id]['status'] = 'stopping'

    return jsonify({'success': True})


@app.route('/api/stitch/server-command', methods=['POST'])
def generate_server_command():
    """
    Generate a command to run batch processing on the remote server (4T).
    Returns the command string that can be copied and run via SSH.

    Request body includes all UI settings: cuts, race_info, logos, etc.
    """
    data = request.get_json() or {}
    race = data.get('race', 'U14 run 1')
    date = data.get('date', '20260201')
    workers = data.get('workers', 16)
    test_count = data.get('test_count', 0)
    generate_comparison = data.get('generate_comparison', False)
    vola_file = data.get('vola_file')
    results_file = data.get('results_file')
    cuts = data.get('cuts', [])
    race_info = data.get('race_info', {})
    selected_logos = data.get('selected_logos', [])
    pre_buffer = data.get('pre_buffer', 2.0)
    post_buffer = data.get('post_buffer', 2.0)

    # Build command
    cmd_parts = [
        'cd /home/pa91/skiframes/photo-montages',
        'source venv/bin/activate',
        'export PATH=/home/pa91/bin:$PATH',  # Use NVENC-enabled ffmpeg
        'export SKIFRAMES_DATA_DIR=/home/pa91/data',
        f'python3 edge/batch_stitch.py --race "{race}" --date {date} --workers {workers}'
    ]

    # Add pre/post buffer settings
    cmd_parts[-1] += f' --pre-buffer {pre_buffer} --post-buffer {post_buffer}'

    if test_count > 0:
        cmd_parts[-1] += f' --test {test_count}'

    if generate_comparison:
        cmd_parts[-1] += ' --comparison'

    if vola_file:
        # Convert local path to server path
        server_vola = vola_file.replace('/Volumes/OWC_48/data', '/home/pa91/data')
        cmd_parts[-1] += f' --vola "{server_vola}"'

    if results_file:
        # Convert local path to server path
        server_results = results_file.replace('/Volumes/OWC_48/data', '/home/pa91/data')
        cmd_parts[-1] += f' --results "{server_results}"'

    # Add cut percentages if provided (for middle cameras - first/last use pre/post buffer)
    if cuts:
        for cut in cuts:
            camera = cut.get('camera', '').lower()
            start = cut.get('start_pct', 0) * 100
            end = cut.get('end_pct', 0) * 100
            if camera == 'r1':
                cmd_parts[-1] += f' --r1-start {start:.0f} --r1-end {end:.0f}'
            elif camera == 'axis':
                cmd_parts[-1] += f' --axis-start {start:.0f} --axis-end {end:.0f}'
            elif camera == 'r2':
                cmd_parts[-1] += f' --r2-start {start:.0f} --r2-end {end:.0f}'
            elif camera == 'r3':
                cmd_parts[-1] += f' --r3-start {start:.0f} --r3-end {end:.0f}'

    # Add race info
    if race_info.get('event'):
        cmd_parts[-1] += f' --event "{race_info["event"]}"'
    if race_info.get('discipline'):
        cmd_parts[-1] += f' --discipline "{race_info["discipline"]}"'

    # Add logos
    if selected_logos:
        cmd_parts[-1] += f' --logos "{",".join(selected_logos)}"'

    full_command = ' && '.join(cmd_parts)

    # Also provide SSH command for easy copy-paste
    # Use single quotes around the whole command to preserve inner double quotes
    ssh_command = f"ssh 4t '{full_command}'"

    return jsonify({
        'command': full_command,
        'ssh_command': ssh_command,
        'server': '4t',
        'race': race,
        'date': date,
        'workers': workers
    })


# ============================================
# GATE CALIBRATION ENDPOINTS
# ============================================

@app.route('/api/calibrate/gate-specs')
def get_gate_specs():
    """Return available gate/discipline presets."""
    return jsonify(GATE_SPECS)


@app.route('/api/calibrate/detect', methods=['POST'])
def detect_gates_endpoint():
    """Run gate detection on an already-grabbed frame."""
    data = request.json or {}
    frame_id = data.get('frame_id')
    discipline = data.get('discipline', 'sl_adult')
    min_height = int(data.get('min_height', 30))
    max_height = int(data.get('max_height', 800))
    sat_thresh = int(data.get('sat_thresh', 120))
    val_thresh = int(data.get('val_thresh', 80))

    if not frame_id:
        return jsonify({'error': 'frame_id is required'}), 400

    frame_path = CALIBRATION_FRAMES_DIR / f'{frame_id}.jpg'
    if not frame_path.exists():
        return jsonify({'error': f'Frame {frame_id} not found'}), 404

    frame = cv2.imread(str(frame_path))
    if frame is None:
        return jsonify({'error': 'Failed to read frame'}), 500

    gates = detect_gates(frame, discipline=discipline, min_height=min_height,
                         max_height=max_height, sat_thresh=sat_thresh,
                         val_thresh=val_thresh)

    # Save annotated frame
    annotated = draw_gates_on_frame(frame, gates)
    annotated_path = CALIBRATION_FRAMES_DIR / f'{frame_id}_gates.jpg'
    cv2.imwrite(str(annotated_path), annotated, [cv2.IMWRITE_JPEG_QUALITY, 90])

    return jsonify({
        'success': True,
        'gates': gates,
        'gate_count': len(gates),
        'discipline': discipline,
        'annotated_frame_url': f'/api/calibrate/frame/{frame_id}/annotated',
    })


@app.route('/api/calibrate/frame/<frame_id>/annotated')
def get_annotated_frame(frame_id):
    """Serve annotated frame with detected gates drawn."""
    # Sanitize frame_id
    safe_id = re.sub(r'[^a-zA-Z0-9_-]', '', frame_id)
    path = CALIBRATION_FRAMES_DIR / f'{safe_id}_gates.jpg'
    if not path.exists():
        return jsonify({'error': 'Annotated frame not found'}), 404
    return send_file(str(path), mimetype='image/jpeg')


@app.route('/api/calibrate/compute', methods=['POST'])
def compute_calibration_endpoint():
    """Compute calibration from coach-adjusted gates."""
    data = request.json or {}
    frame_id = data.get('frame_id')
    camera_id = data.get('camera_id')
    discipline = data.get('discipline', 'sl_adult')
    gates = data.get('gates', [])

    if not frame_id:
        return jsonify({'error': 'frame_id is required'}), 400
    if not camera_id:
        return jsonify({'error': 'camera_id is required'}), 400
    if len(gates) < 4:
        return jsonify({'error': f'Need at least 4 gates, got {len(gates)}'}), 400

    frame_path = CALIBRATION_FRAMES_DIR / f'{frame_id}.jpg'
    if not frame_path.exists():
        return jsonify({'error': f'Frame {frame_id} not found'}), 404

    frame = cv2.imread(str(frame_path))
    if frame is None:
        return jsonify({'error': 'Failed to read frame'}), 500

    gate_spec = GATE_SPECS.get(discipline, GATE_SPECS['sl_adult'])

    # GPS world coordinates (Mode B) — None falls back to Mode A
    world_coords = data.get('world_coords')

    try:
        calibration = compute_calibration(gates, gate_spec, frame.shape, world_coords=world_coords)
    except Exception as e:
        return jsonify({'error': f'Calibration failed: {str(e)}'}), 500

    calibration['camera_id'] = camera_id
    calibration['discipline'] = discipline
    calibration['frame_id'] = frame_id

    # Store GPS metadata if provided
    if data.get('gps_data'):
        calibration['gps_data'] = data['gps_data']
        calibration['calibration_mode'] = 'gps'
    else:
        calibration['calibration_mode'] = 'pixel'

    # Generate calibration ID
    ts = datetime.now().strftime('%Y-%m-%d_%H%M')
    cal_id = f'{camera_id}_{ts}'
    calibration['calibration_id'] = cal_id

    # Draw verification grid
    try:
        verify_frame = draw_verification_grid(frame, calibration)
        verify_path = CALIBRATION_FRAMES_DIR / f'{frame_id}_verify.jpg'
        cv2.imwrite(str(verify_path), verify_frame, [cv2.IMWRITE_JPEG_QUALITY, 90])
    except Exception as e:
        return jsonify({'error': f'Verification grid failed: {str(e)}'}), 500

    # Store pending calibration
    with pending_calibrations_lock:
        pending_calibrations[cal_id] = calibration

    return jsonify({
        'success': True,
        'calibration_id': cal_id,
        'reprojection_error': calibration['reprojection_error'],
        'per_gate_errors': calibration.get('per_gate_errors', []),
        'focal_length': calibration.get('focal_length'),
        'verification_frame_url': f'/api/calibrate/frame/{frame_id}/verification',
        'gate_count': len(gates),
    })


@app.route('/api/calibrate/frame/<frame_id>/verification')
def get_verification_frame(frame_id):
    """Serve verification grid frame."""
    safe_id = re.sub(r'[^a-zA-Z0-9_-]', '', frame_id)
    path = CALIBRATION_FRAMES_DIR / f'{safe_id}_verify.jpg'
    if not path.exists():
        return jsonify({'error': 'Verification frame not found'}), 404
    return send_file(str(path), mimetype='image/jpeg')


@app.route('/api/calibrate/accept', methods=['POST'])
def accept_calibration():
    """Save a computed calibration to disk."""
    data = request.json or {}
    cal_id = data.get('calibration_id')

    if not cal_id:
        return jsonify({'error': 'calibration_id is required'}), 400

    with pending_calibrations_lock:
        calibration = pending_calibrations.pop(cal_id, None)

    if calibration is None:
        return jsonify({'error': f'No pending calibration {cal_id}'}), 404

    # Save calibration JSON
    config_path = CALIBRATION_CONFIG_DIR / f'{cal_id}.json'
    with open(config_path, 'w') as f:
        json.dump(calibration, f, indent=2)

    # Copy reference frame
    frame_id = calibration.get('frame_id', '')
    src_frame = CALIBRATION_FRAMES_DIR / f'{frame_id}.jpg'
    if src_frame.exists():
        import shutil
        dst_frame = CALIBRATION_CONFIG_DIR / f'{cal_id}_frame.jpg'
        shutil.copy2(str(src_frame), str(dst_frame))

    return jsonify({
        'success': True,
        'config_path': str(config_path),
        'calibration_id': cal_id,
    })


@app.route('/api/calibrate/status')
def calibration_status():
    """Return current calibration status for all cameras."""
    cameras = {}
    for cam_id in CAMERAS:
        # Find most recent calibration file for this camera
        pattern = f'{cam_id}_*.json'
        cal_files = sorted(CALIBRATION_CONFIG_DIR.glob(pattern), reverse=True)
        if cal_files:
            with open(cal_files[0]) as f:
                cal = json.load(f)
            cameras[cam_id] = {
                'calibrated': True,
                'calibration_id': cal.get('calibration_id', cal_files[0].stem),
                'timestamp': cal.get('timestamp', ''),
                'reprojection_error': cal.get('reprojection_error', None),
                'discipline': cal.get('discipline', ''),
            }
        else:
            cameras[cam_id] = {'calibrated': False}

    return jsonify({'cameras': cameras})


@app.route('/api/video/upload', methods=['POST'])
def upload_video():
    """Accept a video file upload for testing."""
    if 'video' not in request.files:
        return jsonify({'error': 'No video file in request'}), 400

    video_file = request.files['video']
    if not video_file.filename:
        return jsonify({'error': 'Empty filename'}), 400

    # Size check (500MB limit)
    content_length = request.content_length
    if content_length and content_length > 500 * 1024 * 1024:
        return jsonify({'error': 'File too large (max 500MB)'}), 413

    # Save with UUID prefix to avoid collisions
    safe_name = re.sub(r'[^a-zA-Z0-9._-]', '_', video_file.filename)
    dest_name = f'{uuid.uuid4().hex[:8]}_{safe_name}'
    dest_path = UPLOAD_DIR / dest_name
    video_file.save(str(dest_path))

    # Get video metadata
    cap = cv2.VideoCapture(str(dest_path))
    info = {}
    if cap.isOpened():
        info['fps'] = cap.get(cv2.CAP_PROP_FPS)
        info['frame_count'] = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        info['width'] = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        info['height'] = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        if info['fps'] > 0:
            info['duration'] = info['frame_count'] / info['fps']
        else:
            info['duration'] = 0
        cap.release()

    size_mb = dest_path.stat().st_size / (1024 * 1024)

    return jsonify({
        'success': True,
        'video_path': str(dest_path),
        'filename': video_file.filename,
        'size_mb': round(size_mb, 1),
        **info,
    })


# =============================================================================
# DEVICE INFO & HEARTBEAT
# =============================================================================

@app.route('/api/device/info')
def device_info():
    """Return device identity and status for coach page discovery."""
    active_sessions = []
    with jobs_lock:
        for job_id, job in active_jobs.items():
            if job.get('status') == 'running':
                active_sessions.append({
                    'job_id': job_id,
                    'config_path': os.path.basename(job.get('config_path', '')),
                    'started_at': job.get('started_at', ''),
                })

    return jsonify({
        'device_id': DEVICE_ID,
        'cameras': [{'id': cid, 'name': cam.get('name', cid)}
                     for cid, cam in CAMERAS.items()],
        'active_sessions': active_sessions,
        'hostname': socket.gethostname(),
        'port': 5000,
        'timestamp': datetime.now().isoformat(),
    })


def _device_heartbeat_loop():
    """Background thread: send periodic heartbeats to the admin API."""
    import time
    try:
        import requests as http_req
    except ImportError:
        print(f"[HEARTBEAT] requests not installed, heartbeat disabled")
        return

    while True:
        try:
            active = []
            with jobs_lock:
                for job_id, job in active_jobs.items():
                    if job.get('status') == 'running':
                        # Get session_id from runner config if available
                        sid = ''
                        cfg_path = job.get('config_path', '')
                        if cfg_path:
                            sid = os.path.basename(cfg_path).replace('.json', '')
                        active.append({
                            'job_id': job_id,
                            'session_id': sid,
                            'started_at': job.get('started_at', ''),
                        })
            payload = {
                'device_id': DEVICE_ID,
                'cameras': list(CAMERAS.keys()),
                'active_sessions': active,
                'hostname': socket.gethostname(),
                'timestamp': datetime.now().isoformat(),
            }
            http_req.post(f'{ADMIN_API_URL}/device/heartbeat', json=payload, timeout=5)
        except Exception as e:
            pass  # Silent fail — heartbeat is best-effort
        time.sleep(30)


# Start heartbeat thread
_heartbeat_thread = threading.Thread(target=_device_heartbeat_loop, daemon=True)
_heartbeat_thread.start()


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=os.environ.get('FLASK_DEBUG', '0') == '1')
