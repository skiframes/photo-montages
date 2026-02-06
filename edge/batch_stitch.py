#!/usr/bin/env python3
"""
Batch video stitching script for server-side processing.
Run this on the Tesla GPU server without needing a web UI.

Usage:
    python3 batch_stitch.py --vola /path/to/vola.xlsx --race "U14 run 1" --output /path/to/output

Example on server:
    cd /home/pa91/skiframes/photo-montages
    source venv/bin/activate
    export SKIFRAMES_DATA_DIR=/home/pa91/data
    python3 edge/batch_stitch.py --race "U14 run 1" --workers 16

With racer names from PDF:
    python3 edge/batch_stitch.py --race "U14 run 1" --results /path/to/results.pdf --workers 16
"""

import os
import re
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime

# Use larger temp directory if available (avoid filling up small /tmp partitions)
_alt_tmp = os.path.expanduser('~/tmp')
if os.path.isdir(_alt_tmp) and not os.environ.get('TMPDIR'):
    os.environ['TMPDIR'] = _alt_tmp
    os.environ['TEMP'] = _alt_tmp
    os.environ['TMP'] = _alt_tmp
    # Also set tempfile.tempdir directly to override any cached value
    import tempfile
    tempfile.tempdir = _alt_tmp
    print(f"[batch] Using temp directory: {_alt_tmp}")

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))

from stitcher import VideoStitcher, Racer, CameraCut

# Try to import pdfplumber for PDF parsing
try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

# Data directory from environment
DATA_BASE_DIR = Path(os.environ.get('SKIFRAMES_DATA_DIR', '/Volumes/OWC_48/data'))
VOLA_DIR = DATA_BASE_DIR / 'vola'
RECORDINGS_DIR = DATA_BASE_DIR / 'recordings'


def find_vola_file():
    """Find the most recent Vola Excel file."""
    vola_files = list(VOLA_DIR.glob('**/*.xlsx'))
    # Filter out temp files
    vola_files = [f for f in vola_files if not f.name.startswith('~$')]
    if not vola_files:
        return None
    # Return most recent
    return max(vola_files, key=lambda f: f.stat().st_mtime)


def find_recordings(race_date: str = '20260201'):
    """Find video recordings for each camera.

    Supports two directory structures:
    - Server: /sd_R1/2026-02-01/*.mp4, /sd_Axis/20260201/**/*.mkv
    - Mac: /sd_R1/sdcard/Mp4Record/20260201/*.mp4, /sd_Axis/sdcard/20260201/*.mkv
    """
    video_paths = {}

    # Format date for different path styles
    date_hyphenated = f"{race_date[:4]}-{race_date[4:6]}-{race_date[6:8]}"  # 2026-02-01

    camera_folders = {
        'R1': 'sd_R1',
        'R2': 'sd_R2',
        'R3': 'sd_R3',
        'Axis': 'sd_Axis'
    }

    for camera, folder in camera_folders.items():
        if camera == 'Axis':
            # Try server structure first: /sd_Axis/20260201/**/*.mkv
            axis_path = RECORDINGS_DIR / folder / race_date
            if axis_path.exists():
                videos = sorted(axis_path.glob('**/*.mkv'))
                if videos:
                    video_paths[camera] = [str(v) for v in videos]
            else:
                # Try Mac structure: /sd_Axis/sdcard/20260201/*.mkv
                axis_path = RECORDINGS_DIR / folder / 'sdcard' / race_date
                if axis_path.exists():
                    videos = sorted(axis_path.glob('*.mkv'))
                    if videos:
                        video_paths[camera] = [str(v) for v in videos]
        else:
            # Try server structure first: /sd_R1/2026-02-01/*.mp4
            reolink_path = RECORDINGS_DIR / folder / date_hyphenated
            if reolink_path.exists():
                videos = sorted(reolink_path.glob('*.mp4'))
                if videos:
                    video_paths[camera] = [str(v) for v in videos]
            else:
                # Try Mac structure: /sd_R1/sdcard/Mp4Record/20260201/*.mp4
                reolink_path = RECORDINGS_DIR / folder / 'sdcard' / 'Mp4Record' / race_date
                if reolink_path.exists():
                    videos = sorted(reolink_path.glob('*.mp4'))
                    if videos:
                        video_paths[camera] = [str(v) for v in videos]

    return video_paths


def parse_vola_excel(vola_file: Path, race: str):
    """Parse Vola Excel file to get racer timing data."""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl not installed")
        sys.exit(1)

    # Map race selection to sheet names
    race_map = {
        'u12 run 1': ('u12 start run 1', 'u12 end run 1'),
        'u12 run 2': ('u12 start run 2', 'u12 end run 2'),
        'u14 run 1': ('u14 start run 1', 'u14 end run 1'),
        'u14 run 2': ('u14 start run 2', 'u14 end run 2'),
    }

    race_key = race.lower()
    if race_key not in race_map:
        print(f"Error: Invalid race '{race}'. Must be one of: {list(race_map.keys())}")
        sys.exit(1)

    start_sheet, end_sheet = race_map[race_key]

    wb = openpyxl.load_workbook(vola_file, data_only=True)

    def parse_vola_time(time_str):
        """Parse Vola time format like '11h43:58.5258' to seconds."""
        time_str = str(time_str).strip()
        # Handle format: 11h43:58.5258
        if 'h' in time_str:
            h_part, rest = time_str.split('h')
            m_part, s_part = rest.split(':')
            return int(h_part) * 3600 + int(m_part) * 60 + float(s_part)
        # Handle format: 11:43:58.5258
        parts = time_str.split(':')
        if len(parts) == 3:
            h, m, s = parts
            return int(h) * 3600 + int(m) * 60 + float(s)
        return None

    # Parse start times
    start_times = {}
    ws_start = wb[start_sheet]
    for row in ws_start.iter_rows(min_row=2, values_only=True):
        bib_raw = row[1]
        time_str = row[2]
        if bib_raw and time_str:
            try:
                bib = int(bib_raw)
                seconds = parse_vola_time(time_str)
                if seconds is not None:
                    start_times[bib] = seconds
            except (ValueError, TypeError):
                continue

    # Parse end times
    end_times = {}
    ws_end = wb[end_sheet]
    for row in ws_end.iter_rows(min_row=2, values_only=True):
        bib_raw = row[1]
        time_str = row[2]
        if bib_raw:
            try:
                bib = int(bib_raw)
                if time_str:
                    seconds = parse_vola_time(time_str)
                    if seconds is not None:
                        end_times[bib] = seconds
            except (ValueError, TypeError):
                continue

    wb.close()

    # Build racer list
    racers = []
    default_dnf_duration = 60.0  # Default duration for DSQ/DNF racers (no finish time)
    for bib in sorted(start_times.keys()):
        start_sec = start_times[bib]

        if bib in end_times:
            end_sec = end_times[bib]
            duration = end_sec - start_sec
            if duration <= 0:
                continue
            status = 'finished'
        else:
            # DSQ/DNF/DNS: no finish time, use default duration
            duration = default_dnf_duration
            end_sec = start_sec + duration
            status = 'DNF'

        # Determine gender from bib using NHARA bib ranges
        gender = get_gender_from_bib(bib)

        racers.append(Racer(
            bib=bib,
            name=f'Bib{bib}',  # Will be replaced if startlist is provided
            team='',
            gender=gender,
            start_time_sec=start_sec,
            finish_time_sec=end_sec,
            duration=duration,
            ussa_id='',
            status=status
        ))

    return racers


def get_default_cuts():
    """Default cut percentages for 4-camera setup."""
    return [
        CameraCut(camera='R1', start_pct=-0.05, end_pct=0.05),
        CameraCut(camera='Axis', start_pct=0.05, end_pct=0.55),
        CameraCut(camera='R2', start_pct=0.55, end_pct=0.72),
        CameraCut(camera='R3', start_pct=0.72, end_pct=1.05),
    ]


def get_gender_from_bib(bib: int) -> str:
    """Determine gender from bib number based on NHARA convention."""
    if 1 <= bib <= 60 or 111 <= bib <= 170:
        return 'Women'
    elif 61 <= bib <= 99 or 171 <= bib <= 220:
        return 'Men'
    # Odd = women, even = men as fallback
    return 'Women' if bib % 2 == 1 else 'Men'


def parse_results_pdf(pdf_path: str) -> dict:
    """
    Parse a results PDF to extract racer names, USSA IDs, and status.

    Returns dict mapping bib number to:
    {'name': 'Firstname Lastname', 'team': 'SUN', 'ussa_id': 'E7031024', 'gender': 'Women', 'status': 'finished'}
    """
    if not HAS_PDFPLUMBER:
        print("Warning: pdfplumber not installed, cannot parse results PDF")
        return {}

    bib_to_racer = {}
    current_gender = None
    current_status = 'finished'

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                for line in text.split('\n'):
                    # Detect gender sections
                    if 'Women' in line and len(line.strip()) < 20:
                        current_gender = 'Women'
                        continue
                    if 'Men' in line and len(line.strip()) < 20:
                        current_gender = 'Men'
                        continue

                    # Detect status sections
                    if 'Did Not Start' in line:
                        current_status = 'DNS'
                        continue
                    if 'Did Not Finish' in line:
                        current_status = 'DNF'
                        continue
                    if 'Disqualified' in line:
                        current_status = 'DSQ'
                        continue

                    # Parse ranked finishers
                    # Pattern: rank bib [ussa_id] lastname firstname year club run1 run2 total
                    ranked_match = re.match(
                        r'(\d+)\s+(\d+)\s+(?:(E\d{7})\s+)?((?:[A-Za-z\'\-]+\s+)+)(\d{4})\s+([A-Z]{2,4})\s+([\d:.]+)\s+([\d:.]+)\s+([\d:.]+)',
                        line
                    )
                    if ranked_match:
                        bib = int(ranked_match.group(2))
                        ussa_id = ranked_match.group(3) or ''
                        name_parts = ranked_match.group(4).strip().split()
                        team = ranked_match.group(6)

                        # Name: last word is firstname
                        if len(name_parts) >= 2:
                            firstname = name_parts[-1]
                            lastname = ' '.join(name_parts[:-1])
                        else:
                            firstname = name_parts[0] if name_parts else ''
                            lastname = ''

                        bib_to_racer[bib] = {
                            'name': f"{firstname} {lastname}".strip(),
                            'team': team,
                            'ussa_id': ussa_id,
                            'gender': current_gender or get_gender_from_bib(bib),
                            'status': 'finished'
                        }
                        continue

                    # Parse DSQ/DNF/DNS entries
                    unranked_match = re.match(
                        r'(\d+)\s+(?:(E\d{7})\s+)?((?:[A-Za-z\'\-]+\s+)+)(\d{4})\s+([A-Z]{2,4})',
                        line
                    )
                    if unranked_match and current_status != 'finished':
                        bib = int(unranked_match.group(1))
                        ussa_id = unranked_match.group(2) or ''
                        name_parts = unranked_match.group(3).strip().split()
                        team = unranked_match.group(5)

                        if len(name_parts) >= 2:
                            firstname = name_parts[-1]
                            lastname = ' '.join(name_parts[:-1])
                        else:
                            firstname = name_parts[0] if name_parts else ''
                            lastname = ''

                        bib_to_racer[bib] = {
                            'name': f"{firstname} {lastname}".strip(),
                            'team': team,
                            'ussa_id': ussa_id,
                            'gender': current_gender or get_gender_from_bib(bib),
                            'status': current_status
                        }
        # Fallback: try table extraction for bibs missed by text parsing
        # Some PDF rows overlap with page headers, garbling text extraction
        # but table extraction can still find the cell contents
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row:
                            continue
                        # Find bib number in row cells
                        bib = None
                        ussa_id = ''
                        name_cell = ''
                        team = ''
                        for cell in row:
                            if cell is None:
                                continue
                            cell = str(cell).strip()
                            # Look for bib number (3 digit, in range 100-220)
                            if not bib and re.match(r'^\d{3}$', cell):
                                candidate = int(cell)
                                if 100 <= candidate <= 220 and candidate not in bib_to_racer:
                                    bib = candidate
                            # Also check cells with newlines (garbled with headers)
                            if not bib:
                                bib_in_cell = re.search(r'\b(1[1-9]\d|20\d|21\d|220)\b', cell)
                                if bib_in_cell:
                                    candidate = int(bib_in_cell.group(1))
                                    if candidate not in bib_to_racer:
                                        bib = candidate
                            # Look for USSA ID
                            ussa_match = re.search(r'(E\d{7})', cell)
                            if ussa_match:
                                ussa_id = ussa_match.group(1)
                            # Look for name (Lastname Firstname pattern)
                            name_match = re.search(r'([A-Z][a-z]+(?:[\'\\-][A-Z][a-z]+)?\s+[A-Z][a-z]+)', cell)
                            if name_match and not name_cell:
                                name_cell = name_match.group(1)
                            # Look for team (2-4 uppercase letters alone)
                            team_match = re.search(r'\b([A-Z]{2,4})\b', cell)
                            if team_match and team_match.group(1) not in ('None', 'THE', 'AND', 'FOR'):
                                # Only set if looks like a team code
                                candidate_team = team_match.group(1)
                                if candidate_team in ('SUN', 'PROC', 'RMS', 'FS', 'CMS'):
                                    team = candidate_team

                        if bib and bib not in bib_to_racer and (ussa_id or name_cell):
                            # Parse name: "Lastname Firstname" -> "Firstname Lastname"
                            name = ''
                            if name_cell:
                                parts = name_cell.split()
                                if len(parts) >= 2:
                                    name = f"{parts[-1]} {' '.join(parts[:-1])}"
                                else:
                                    name = name_cell

                            bib_to_racer[bib] = {
                                'name': name,  # Leave empty if not found; startlist can fill it
                                'team': team,
                                'ussa_id': ussa_id,
                                'gender': get_gender_from_bib(bib),
                                'status': 'finished'  # Can't determine from garbled table
                            }
                            print(f"  Table fallback: Bib {bib} -> {name or '(no name)'} ({team}) ussa={ussa_id}")

    except Exception as e:
        print(f"Error parsing results PDF {pdf_path}: {e}")

    return bib_to_racer


def parse_startlist_pdf(pdf_path: str) -> dict:
    """
    Parse a start list PDF to extract bib-to-racer mapping.

    Start list format: "111 Stellato Brigid 2013 SUN"
    (bib lastname firstname year team)

    Returns dict mapping bib number to {'name': 'Firstname Lastname', 'team': 'SUN', 'gender': 'Women'}
    """
    if not HAS_PDFPLUMBER:
        return {}

    bib_to_racer = {}

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                for line in text.split('\n'):
                    matches = re.findall(r'(\d+)\s+((?:[A-Za-z\'\-]+\s+){1,3}[A-Za-z\'\-]+)\s+(\d{4})\s+([A-Z]{2,4})', line)
                    for match in matches:
                        bib = int(match[0])
                        name_parts = match[1].strip().split()
                        team = match[3]

                        if len(name_parts) >= 2:
                            firstname = name_parts[-1]
                            lastname = ' '.join(name_parts[:-1])
                        else:
                            firstname = name_parts[0] if name_parts else ''
                            lastname = ''

                        bib_to_racer[bib] = {
                            'name': f"{firstname} {lastname}".strip(),
                            'team': team,
                            'gender': get_gender_from_bib(bib)
                        }
    except Exception as e:
        print(f"Error parsing start list PDF {pdf_path}: {e}")

    return bib_to_racer


def find_startlist_pdf(race: str = None):
    """Find a start list PDF in the vola directory."""
    pdf_files = list(VOLA_DIR.glob('**/*.pdf'))
    startlist_files = [f for f in pdf_files if 'start' in f.name.lower()]

    if race and startlist_files:
        race_lower = race.lower()
        # Match run number
        if 'run 1' in race_lower:
            run1_files = [f for f in startlist_files if 'run-1' in f.name.lower() or 'run1' in f.name.lower()]
            if run1_files:
                return max(run1_files, key=lambda f: f.stat().st_mtime)
        elif 'run 2' in race_lower:
            run2_files = [f for f in startlist_files if 'run-2' in f.name.lower() or 'run2' in f.name.lower()]
            if run2_files:
                return max(run2_files, key=lambda f: f.stat().st_mtime)

    if startlist_files:
        return max(startlist_files, key=lambda f: f.stat().st_mtime)
    return None


def find_results_pdf(race: str = None):
    """Find a results PDF in the vola directory matching the race."""
    pdf_files = list(VOLA_DIR.glob('**/*.pdf'))
    # Filter for results files
    results_files = [f for f in pdf_files if 'result' in f.name.lower()]

    if race and results_files:
        # Try to match the race age group (U12 or U14)
        race_lower = race.lower()
        if 'u14' in race_lower:
            u14_files = [f for f in results_files if 'u14' in f.name.lower()]
            if u14_files:
                return max(u14_files, key=lambda f: f.stat().st_mtime)
        elif 'u12' in race_lower:
            u12_files = [f for f in results_files if 'u12' in f.name.lower()]
            if u12_files:
                return max(u12_files, key=lambda f: f.stat().st_mtime)

    if results_files:
        return max(results_files, key=lambda f: f.stat().st_mtime)
    # Return any PDF
    if pdf_files:
        return max(pdf_files, key=lambda f: f.stat().st_mtime)
    return None


def main():
    parser = argparse.ArgumentParser(description='Batch video stitching for ski race footage')
    parser.add_argument('--config', type=str, help='Path to JSON config file (exported from UI)')
    parser.add_argument('--vola', type=str, help='Path to Vola Excel file (auto-detected if not specified)')
    parser.add_argument('--results', type=str, help='Path to results PDF for racer names (auto-detected if not specified)')
    parser.add_argument('--race', type=str, help='Race to process (e.g., "U14 run 1")')
    parser.add_argument('--date', type=str, default='20260201', help='Race date in YYYYMMDD format')
    parser.add_argument('--output', type=str, default='./output', help='Output directory')
    parser.add_argument('--workers', type=int, default=16, help='Number of parallel workers')
    parser.add_argument('--test', type=int, default=0, help='Only process first N racers (0 = all)')
    parser.add_argument('--comparison', action='store_true', help='Generate comparison videos vs fastest')
    parser.add_argument('--list-only', action='store_true', help='List racers without processing')
    parser.add_argument('--bibs', type=str, help='Comma-separated list of bib numbers to process (e.g., "181,182,187")')
    # Pre/post buffer in seconds (for first and last camera)
    parser.add_argument('--pre-buffer', type=float, default=2.0, help='Seconds before racer start (default: 2.0)')
    parser.add_argument('--post-buffer', type=float, default=2.0, help='Seconds after racer finish (default: 2.0)')
    # Camera cut percentages (for middle cameras)
    parser.add_argument('--r1-start', type=float, default=0, help='R1 start percentage (default: 0, uses pre-buffer)')
    parser.add_argument('--r1-end', type=float, default=5, help='R1 end percentage (default: 5)')
    parser.add_argument('--axis-start', type=float, default=5, help='Axis start percentage (default: 5)')
    parser.add_argument('--axis-end', type=float, default=55, help='Axis end percentage (default: 55)')
    parser.add_argument('--r2-start', type=float, default=55, help='R2 start percentage (default: 55)')
    parser.add_argument('--r2-end', type=float, default=72, help='R2 end percentage (default: 72)')
    parser.add_argument('--r3-start', type=float, default=72, help='R3 start percentage (default: 72)')
    parser.add_argument('--r3-end', type=float, default=100, help='R3 end percentage (default: 100, uses post-buffer)')
    # Race info
    parser.add_argument('--event', type=str, default='Western Division Ranking', help='Event name')
    parser.add_argument('--discipline', type=str, default='SL', help='Discipline (SL, GS, etc.)')
    # Logos
    parser.add_argument('--logos', type=str, help='Comma-separated list of logo filenames (e.g., "nhara.png,rmst.png")')

    args = parser.parse_args()

    # Load config file if provided
    config = {}
    if args.config:
        config_path = Path(args.config)
        if config_path.exists():
            with open(config_path) as f:
                config = json.load(f)
            print(f"Loaded config from: {config_path}")
        else:
            print(f"Warning: Config file not found: {config_path}")

    # Config file values override defaults, command line overrides config
    def get_setting(name, default, config_key=None):
        """Get setting with priority: command line > config file > default"""
        cli_val = getattr(args, name, None)
        if cli_val is not None and cli_val != default:
            return cli_val
        if config_key and config_key in config:
            return config[config_key]
        if name in config:
            return config[name]
        return default if cli_val is None else cli_val

    # Get race from args or config
    race = args.race or config.get('race')
    if not race:
        print("Error: --race is required (or provide via config file)")
        sys.exit(1)

    print("=" * 60)
    print("Skiframes Batch Video Stitcher")
    print("=" * 60)
    print(f"Data directory: {DATA_BASE_DIR}")
    print(f"Race: {race}")
    print(f"Date: {args.date}")
    print(f"Workers: {args.workers}")
    print()

    # Find Vola file
    if args.vola:
        vola_file = Path(args.vola)
    else:
        vola_file = find_vola_file()

    if not vola_file or not vola_file.exists():
        print(f"Error: Vola file not found")
        print(f"Searched in: {VOLA_DIR}")
        sys.exit(1)

    # Find results PDF for racer names
    results_pdf = None
    if args.results:
        results_pdf = Path(args.results)
    else:
        results_pdf = find_results_pdf(race)

    bib_to_info = {}
    if results_pdf and results_pdf.exists():
        print(f"Results PDF: {results_pdf}")
        bib_to_info = parse_results_pdf(str(results_pdf))
        print(f"  Found {len(bib_to_info)} racers with names")
    else:
        print("Warning: No results PDF found, using bib numbers as names")

    print(f"Vola file: {vola_file}")

    # Parse racers
    print("Parsing racer timing data...")
    racers = parse_vola_excel(vola_file, race)
    print(f"Found {len(racers)} racers with valid timing")

    # Also find start list PDF for supplementary name data
    startlist_pdf = find_startlist_pdf(race)
    startlist_info = {}
    if startlist_pdf and startlist_pdf.exists():
        print(f"Start list PDF: {startlist_pdf}")
        startlist_info = parse_startlist_pdf(str(startlist_pdf))
        print(f"  Found {len(startlist_info)} racers in start list")

    # Merge in names: results PDF takes priority, startlist fills gaps
    for racer in racers:
        info = bib_to_info.get(racer.bib, {})
        sl_info = startlist_info.get(racer.bib, {})

        # Name: prefer results, fallback to startlist
        name = info.get('name') or sl_info.get('name')
        if name:
            racer.name = name
        # Team: prefer results, fallback to startlist
        team = info.get('team') or sl_info.get('team')
        if team:
            racer.team = team
        # USSA ID: only from results PDF
        if info.get('ussa_id'):
            racer.ussa_id = info['ussa_id']
        # Gender: prefer results, fallback to startlist, fallback to bib range
        gender = info.get('gender') or sl_info.get('gender')
        if gender:
            racer.gender = gender
        # Status: update from results PDF (DSQ/DNF/DNS) if available
        if info.get('status'):
            racer.status = info['status']

    # Remove DNS racers (Did Not Start - no video to process)
    dns_racers = [r for r in racers if r.status == 'DNS']
    if dns_racers:
        print(f"  Removing {len(dns_racers)} DNS racers: {[r.bib for r in dns_racers]}")
        racers = [r for r in racers if r.status != 'DNS']

    # Report any racers still without names
    unnamed = [r for r in racers if r.name.startswith('Bib')]
    if unnamed:
        print(f"  Warning: {len(unnamed)} racers still without names: {[r.bib for r in unnamed]}")

    if args.list_only:
        print("\nRacers:")
        for r in racers:
            status_str = f" [{r.status}]" if r.status != 'finished' else ""
            print(f"  Bib {r.bib}: {r.name} ({r.team}) - {r.duration:.2f}s ({r.gender}){status_str}")
        sys.exit(0)

    # Find video recordings
    print("\nSearching for video recordings...")
    video_paths = find_recordings(args.date)

    for camera, paths in video_paths.items():
        print(f"  {camera}: {len(paths)} videos")

    missing = [c for c in ['R1', 'R2', 'R3', 'Axis'] if c not in video_paths]
    if missing:
        print(f"\nWarning: Missing videos for: {missing}")
        response = input("Continue anyway? [y/N] ")
        if response.lower() != 'y':
            sys.exit(1)

    # Filter to specific bibs if requested
    if args.bibs:
        bib_filter = set(int(b.strip()) for b in args.bibs.split(','))
        racers = [r for r in racers if r.bib in bib_filter]
        print(f"\nFiltering to {len(racers)} racers with bibs: {sorted(bib_filter)}")
        if not racers:
            print("ERROR: No matching racers found for specified bibs")
            sys.exit(1)

    # Limit racers if test mode
    if args.test > 0:
        racers = racers[:args.test]
        print(f"\nTest mode: processing {len(racers)} racers")

    # Create output directory
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Get cuts from config or args
    if 'cuts' in config:
        # Config file has cuts as list of dicts
        cuts = [CameraCut(camera=c['camera'], start_pct=c['start_pct'], end_pct=c['end_pct'])
                for c in config['cuts']]
    else:
        # Use command line args
        cuts = [
            CameraCut(camera='R1', start_pct=args.r1_start / 100, end_pct=args.r1_end / 100),
            CameraCut(camera='Axis', start_pct=args.axis_start / 100, end_pct=args.axis_end / 100),
            CameraCut(camera='R2', start_pct=args.r2_start / 100, end_pct=args.r2_end / 100),
            CameraCut(camera='R3', start_pct=args.r3_start / 100, end_pct=args.r3_end / 100),
        ]

    # Race info for overlay - from config or args
    race_info = config.get('race_info', {})
    if not race_info:
        race_info = {
            'event': args.event,
            'discipline': args.discipline,
            'age_group': 'U14' if 'u14' in race.lower() else 'U12',
            'run': 'Run 1' if 'run 1' in race.lower() else 'Run 2',
            'date': f"{args.date[:4]}-{args.date[4:6]}-{args.date[6:8]}",
        }

    # Get logos from config or args
    selected_logos = config.get('selected_logos', [])
    if args.logos:
        selected_logos = [l.strip() for l in args.logos.split(',')]

    print(f"\nStarting video stitching with {args.workers} workers...")
    print(f"Output: {output_dir}")

    # Show cuts, indicating skipped cameras
    cuts_info = []
    for cut in cuts:
        if cut.start_pct == cut.end_pct:
            cuts_info.append(f"{cut.camera}=SKIP")
        else:
            cuts_info.append(f"{cut.camera}={cut.start_pct*100:.0f}%-{cut.end_pct*100:.0f}%")
    print(f"Cuts: {', '.join(cuts_info)}")

    if selected_logos:
        print(f"Logos: {', '.join(selected_logos)}")
    print()

    # Get pre/post buffer settings
    pre_buffer = getattr(args, 'pre_buffer', 2.0)
    post_buffer = getattr(args, 'post_buffer', 2.0)
    print(f"Pre-buffer: {pre_buffer}s, Post-buffer: {post_buffer}s")

    # Create stitcher and run
    stitcher = VideoStitcher(
        racers=racers,
        cuts=cuts,
        video_paths=video_paths,
        output_dir=str(output_dir),
        race_name=race.lower().replace(' ', '_'),
        race_info=race_info,
        selected_logos=selected_logos if selected_logos else None,
        pre_buffer_sec=pre_buffer,
        post_buffer_sec=post_buffer,
    )

    if args.comparison:
        outputs = stitcher.process_all_with_comparison_parallel(
            max_workers=args.workers,
            generate_comparison=True
        )
    else:
        outputs = stitcher.process_all_parallel(max_workers=args.workers)

    print()
    print("=" * 60)
    print(f"Complete! Generated {len(outputs)} videos")
    print(f"Output directory: {output_dir}")
    print("=" * 60)


if __name__ == '__main__':
    main()
